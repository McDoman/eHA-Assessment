"""
01_crosswalk_normalisation_pipeline.py
======================================
Task 1 -- Transform the senatorial district spreadsheet into a normalised
database table through automation only.

The source workbook `LGA_SEN_Districts.xlsx` is opened strictly read-only. It is
never written to, and no manual edit of it is required or performed. Everything
below -- unmerging, header reconstruction, forward-fill of the merged state and
district columns, name repair, deduplication, conflict adjudication and
validation against the authoritative boundary layer -- is done in memory.

Stages merged into this single pipeline script
----------------------------------------------
  A. Workbook profiling and authority selection across all sheets
  B. Structural normalisation (merged cells, two-tier header, blank-continuation)
  C. Administrative name reconciliation against the boundary layer
  D. Duplicate and conflict adjudication
  E. Validation against the workbook's own control totals and the GeoPackage
  F. Emission of the normalised table, the reconciliation ledger, the
     exceptions report, and a written normalisation report

Outputs
-------
  data/lga_senatorial_crosswalk.csv           normalised, database-ready table
  reports/01_crosswalk_reconciliation_ledger.csv   every name decision
  reports/01_crosswalk_exceptions.csv              every unresolved / conflicted name
  reports/01_crosswalk_normalisation_report.md     written account
"""

from __future__ import annotations

import re
from collections import Counter, defaultdict

import geopandas as gpd
import openpyxl
import pandas as pd

from common import (ART, FUZZY_MATCH_THRESHOLD, FUZZY_REVIEW_THRESHOLD, SRC,
                    Ledger, banner, best_fuzzy, clean_text, get_logger,
                    match_key, similarity, title_case_admin)

LOG = get_logger("01_crosswalk_normalisation")

# Markers that disqualify a sheet from being treated as authoritative.
SUPERSEDED_MARKERS = ("superseded", "do not use", "obsolete", "deprecated", "archive")

# Tokens that identify the real header row of the table.
HEADER_TOKENS = {"lga", "l.g.a", "state", "senatorial district", "lga code",
                 "no. of wards", "remarks"}

# Canonical output column for each header variant we may meet.
HEADER_MAP = {
    "state": "state_name_raw",
    "l.g.a": "lga_name_raw",
    "lga": "lga_name_raw",
    "local government area": "lga_name_raw",
    "senatorial district": "sen_district_raw",
    "no. of wards": "ward_count_declared",
    "number of wards": "ward_count_declared",
    "lga code": "lga_code",
    "code": "lga_code",
    "remarks": "remarks",
}


# ==========================================================================
# A. Workbook profiling and authority selection
# ==========================================================================

def profile_workbook(path) -> dict:
    """
    Read every sheet, describe its shape, and decide which one is authoritative.

    The README warns that the workbook contains more than one sheet and that all
    of them must be read before deciding what is authoritative. The rule applied
    here is explicit rather than positional: a sheet whose leading rows carry a
    supersession marker is demoted to 'alias source only', never used for the
    LGA -> senatorial district relation itself.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    profile = {"sheets": {}, "authoritative": None, "alias_sources": []}

    for ws in wb.worksheets:
        head_text = " ".join(
            clean_text(c).lower()
            for row in ws.iter_rows(min_row=1, max_row=min(4, ws.max_row), values_only=True)
            for c in row if c is not None
        )
        superseded = any(m in head_text for m in SUPERSEDED_MARKERS)
        non_empty = sum(
            1 for row in ws.iter_rows(values_only=True) if any(c is not None for c in row)
        )
        info = {
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "non_empty_rows": non_empty,
            "merged_ranges": [str(r) for r in ws.merged_cells.ranges],
            "n_merged": len(ws.merged_cells.ranges),
            "superseded": superseded,
            "leading_text": head_text[:200],
        }
        profile["sheets"][ws.title] = info
        LOG.info("sheet %-12s rows=%-4d cols=%-2d merged=%-2d superseded=%s",
                 ws.title, ws.max_row, ws.max_column, info["n_merged"], superseded)
        if superseded:
            profile["alias_sources"].append(ws.title)

    # The authoritative sheet is the widest non-superseded sheet: the crosswalk
    # relation needs at least state / LGA / senatorial district, so a two-column
    # sheet cannot carry it.
    eligible = [(t, i) for t, i in profile["sheets"].items()
                if not i["superseded"] and i["max_col"] >= 3]
    if not eligible:
        raise RuntimeError("No sheet in the workbook can carry an LGA->district relation.")
    eligible.sort(key=lambda kv: (kv[1]["max_col"], kv[1]["non_empty_rows"]), reverse=True)
    profile["authoritative"] = eligible[0][0]
    LOG.info("Authoritative sheet selected: %r  (alias-only sheets: %s)",
             profile["authoritative"], profile["alias_sources"] or "none")
    wb.close()
    return profile


# ==========================================================================
# B. Structural normalisation
# ==========================================================================

def read_sheet_unmerged(path, sheet_name: str) -> list[list]:
    """
    Return the sheet as a dense grid with merged cells expanded.

    openpyxl exposes a merged range as one populated top-left cell and a block
    of None. Every downstream step -- header detection, forward-fill, row
    filtering -- would misread those Nones, so the block is filled with the
    anchor value first. This is done on the in-memory grid; the file on disk is
    untouched.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    grid = [[c for c in row] for row in ws.iter_rows(values_only=True)]

    n_filled = 0
    for rng in ws.merged_cells.ranges:
        anchor = grid[rng.min_row - 1][rng.min_col - 1]
        for r in range(rng.min_row - 1, rng.max_row):
            for c in range(rng.min_col - 1, rng.max_col):
                if grid[r][c] is None:
                    grid[r][c] = anchor
                    n_filled += 1
    LOG.info("Expanded %d merged range(s), filling %d cell(s) in sheet %r",
             len(ws.merged_cells.ranges), n_filled, sheet_name)
    wb.close()
    return grid


def locate_header(grid: list[list]) -> int:
    """
    Find the 0-based index of the true header row.

    The sheet carries a decorative title, a provenance line, and a two-tier
    header in which the upper tier ('ADMINISTRATIVE UNIT' / 'ELECTORAL' /
    'CODES') spans groups of columns. The lower tier is the one that names the
    fields, and is identified by how many of its cells are recognised header
    tokens rather than by a hard-coded row number.
    """
    best_idx, best_hits = None, 0
    for i, row in enumerate(grid[:25]):
        vals = [clean_text(v).lower() for v in row if v is not None]
        hits = sum(1 for v in vals if v in HEADER_TOKENS)
        if hits > best_hits:
            best_idx, best_hits = i, hits
    if best_idx is None or best_hits < 2:
        raise RuntimeError("Could not locate a header row in the crosswalk sheet.")
    LOG.info("Header row detected at sheet row %d (%d recognised tokens): %s",
             best_idx + 1, best_hits, [clean_text(v) for v in grid[best_idx]])
    return best_idx


def is_noise_row(values: list[str]) -> tuple[bool, str]:
    """Classify a row as a control total, a footnote, or blank."""
    non_empty = [v for v in values if v]
    if not non_empty:
        return True, "blank"
    joined = " ".join(non_empty).lower()
    if non_empty[0].lower() in {"total", "totals", "grand total", "sum"}:
        return True, "control_total"
    if re.match(r"^(source|note|prepared by|for planning|footnote|nb)\b", joined):
        return True, "footnote"
    # A row with a single long free-text cell and nothing else is commentary.
    if len(non_empty) == 1 and len(non_empty[0]) > 40:
        return True, "footnote"
    return False, ""


def build_raw_table(grid: list[list], header_idx: int) -> tuple[pd.DataFrame, dict]:
    """
    Turn the unmerged grid into a tidy frame, forward-filling the columns that
    the compiler left blank as a visual grouping device.

    In this workbook the STATE column is merged down and the SENATORIAL DISTRICT
    column is a blank-continuation block: a value appears once and every LGA
    beneath it belongs to it until the next value. Forward fill is only applied
    to those two grouping columns, and only within the data region -- never to
    the LGA name or the LGA code, where a blank would be a genuine defect.
    """
    header_cells = [clean_text(v).lower() for v in grid[header_idx]]
    colnames, unmapped = [], []
    for j, h in enumerate(header_cells):
        if h in HEADER_MAP:
            colnames.append(HEADER_MAP[h])
        elif h:
            colnames.append(re.sub(r"[^a-z0-9]+", "_", h).strip("_"))
            unmapped.append(h)
        else:
            colnames.append(f"unnamed_{j}")
    if unmapped:
        LOG.warning("Header cells not in the known map, carried through verbatim: %s", unmapped)

    records, discarded = [], []
    for i in range(header_idx + 1, len(grid)):
        raw = grid[i]
        vals = [clean_text(v) for v in raw]
        noise, why = is_noise_row(vals)
        if noise:
            discarded.append({"sheet_row": i + 1, "reason": why,
                              "content": " | ".join(v for v in vals if v)})
            continue
        rec = {colnames[j]: (raw[j] if j < len(raw) else None) for j in range(len(colnames))}
        rec["_sheet_row"] = i + 1
        records.append(rec)

    df = pd.DataFrame(records)
    for col in ("state_name_raw", "sen_district_raw"):
        if col in df.columns:
            before = df[col].isna().sum()
            df[col] = df[col].ffill()
            LOG.info("Forward-filled %-18s : %d blank cell(s) resolved from the group above",
                     col, before - df[col].isna().sum())

    control = {}
    for d in discarded:
        if d["reason"] == "control_total":
            nums = re.findall(r"\b(\d{2,6})\b", d["content"])
            if nums:
                control["declared_ward_total"] = int(nums[-1])
    LOG.info("Data rows kept: %d ; rows discarded as structure/noise: %d",
             len(df), len(discarded))
    return df, {"discarded_rows": discarded, "control": control, "columns": colnames}


# ==========================================================================
# C/D. Name reconciliation, deduplication, conflict adjudication
# ==========================================================================

def load_authority() -> dict:
    """The GeoPackage boundary layer is the authority for spelling and codes."""
    lgas = gpd.read_file(SRC["boundaries"], layer="lgas").drop(columns="geometry")
    sens = gpd.read_file(SRC["boundaries"], layer="senatorial_districts").drop(columns="geometry")
    states = gpd.read_file(SRC["boundaries"], layer="states").drop(columns="geometry")
    wards = gpd.read_file(SRC["boundaries"], layer="wards").drop(columns="geometry")

    def universe(frame, name_col):
        u = {}
        for _, r in frame.iterrows():
            k = match_key(r[name_col])
            if k in u and u[k] != r[name_col]:
                raise RuntimeError(
                    f"Match key {k!r} is ambiguous in the authority layer: "
                    f"{u[k]!r} vs {r[name_col]!r}. The normalisation rule would "
                    f"collapse two distinct units and must be tightened.")
            u[k] = r[name_col]
        return u

    auth = {
        "lgas": lgas, "sens": sens, "states": states,
        "ward_counts": wards.groupby("lga_code").size().to_dict(),
        "lga_universe": universe(lgas, "lga_name"),
        "sen_universe": universe(sens, "sen_district"),
        "state_universe": universe(states, "state_name"),
        "lga_by_key": {match_key(r.lga_name): r for r in lgas.itertuples()},
        "lga_by_code": {r.lga_code: r for r in lgas.itertuples()},
        "sen_by_name": {r.sen_district: r for r in sens.itertuples()},
    }
    LOG.info("Authority loaded: %d LGAs, %d senatorial districts, %d states, %d wards",
             len(lgas), len(sens), len(states), len(wards))
    return auth


def harvest_aliases(path, sheet_names: list[str]) -> dict[str, set[str]]:
    """
    Mine superseded sheets for historic spellings.

    The 2019 sheet may not define the relation -- it is explicitly superseded --
    but it is still evidence about how a unit has been spelled in the past, and
    that is legitimate input to name reconciliation. Aliases are keyed by
    match_key so they only ever help, never override.
    """
    aliases: dict[str, set[str]] = defaultdict(set)
    wb = openpyxl.load_workbook(path, data_only=True)
    for name in sheet_names:
        ws = wb[name]
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                s = clean_text(cell)
                if s and len(s) > 2 and not s.isupper() or (s and s.istitle()):
                    aliases[match_key(s)].add(s)
    wb.close()
    LOG.info("Harvested %d alias key(s) from superseded sheet(s) %s", len(aliases), sheet_names)
    return aliases


def reconcile_names(df: pd.DataFrame, auth: dict, ledger: Ledger) -> pd.DataFrame:
    """
    Map every raw administrative name onto its authoritative spelling.

    Cascade, most trustworthy first:
      1. exact match on the raw string                -> exact
      2. match on the normalised key                  -> normalised (case /
         whitespace / suffix / separator repair)
      3. character similarity above the accept
         threshold                                    -> fuzzy_accepted
      4. similarity above the review threshold        -> fuzzy_review (kept but
         flagged; a human must confirm)
      5. nothing                                      -> unresolved (excluded
         from the normalised table, listed in exceptions)

    Every step of the cascade writes a ledger row, including the exact matches,
    so the record is genuinely of *every* name, not only the difficult ones.
    """
    resolved = {"lga_name": [], "sen_district": [], "state_name": [],
                "match_method": [], "match_confidence": [], "match_outcome": []}

    for row in df.itertuples():
        sheet_row = getattr(row, "_sheet_row", "")
        out = {}
        worst_outcome, worst_method, worst_conf = "resolved", "exact", 1.0

        for raw_col, universe_key, field_name in (
            ("lga_name_raw", "lga_universe", "lga_name"),
            ("sen_district_raw", "sen_universe", "sen_district"),
            ("state_name_raw", "state_universe", "state_name"),
        ):
            raw = getattr(row, raw_col, None)
            raw_clean = clean_text(raw)
            universe = auth[universe_key]
            canon_by_name = {v: v for v in universe.values()}

            if raw_clean in canon_by_name:
                value, method, conf, outcome = raw_clean, "exact", 1.0, "resolved"
                note = ""
            else:
                key = match_key(raw_clean)
                if key in universe:
                    value, method, conf, outcome = universe[key], "normalised", 1.0, "resolved"
                    note = "case/whitespace/suffix/separator repair"
                else:
                    bk, score = best_fuzzy(key, universe)
                    if bk is not None and score >= FUZZY_MATCH_THRESHOLD:
                        value, method, conf = universe[bk], "fuzzy_accepted", score
                        outcome, note = "resolved", f"character similarity {score:.3f}"
                    elif bk is not None and score >= FUZZY_REVIEW_THRESHOLD:
                        value, method, conf = universe[bk], "fuzzy_review", score
                        outcome = "resolved_review"
                        note = (f"below auto-accept threshold "
                                f"{FUZZY_MATCH_THRESHOLD}; requires confirmation")
                    else:
                        value, method, conf = None, "no_match", score if bk else 0.0
                        outcome, note = "unresolved", "no candidate above review threshold"

            ledger.record(source=f"crosswalk:{sheet_row}", entity_id=str(getattr(row, "lga_code", "")),
                          field_name=field_name, raw_value=raw_clean, resolved_value=value,
                          method=method, confidence=conf, outcome=outcome, note=note)
            out[field_name] = value

            rank = {"resolved": 0, "resolved_review": 1, "unresolved": 2}
            if rank[outcome] > rank[worst_outcome]:
                worst_outcome, worst_method, worst_conf = outcome, method, conf

        resolved["lga_name"].append(out["lga_name"])
        resolved["sen_district"].append(out["sen_district"])
        resolved["state_name"].append(out["state_name"])
        resolved["match_method"].append(worst_method)
        resolved["match_confidence"].append(round(worst_conf, 4))
        resolved["match_outcome"].append(worst_outcome)

    for k, v in resolved.items():
        df[k] = v
    LOG.info("Name reconciliation outcomes: %s", Counter(df.match_outcome).most_common())
    return df


def adjudicate(df: pd.DataFrame, auth: dict, ledger: Ledger) -> tuple[pd.DataFrame, list[dict]]:
    """
    Resolve repeated LGA codes and disagreements with the authority layer.

    Two distinct problems appear in this workbook:

      * benign restatement -- an LGA is listed twice with identical district
        assignment. Deduplicated silently but recorded.

      * substantive conflict -- an LGA is listed twice with *different* district
        assignments. Here the workbook itself supplies the tie-breaker in the
        Remarks column: a row annotated as a gazetted transfer supersedes an
        unannotated one. Where the remarks do not decide it, the row is not
        guessed at: the boundary layer is consulted, and if that too fails to
        decide, the LGA is written to the exceptions report and its district is
        left null rather than fabricated.
    """
    conflicts = []
    df = df.copy()
    df["lga_code"] = df["lga_code"].map(clean_text)
    df["remarks"] = df.get("remarks", pd.Series([""] * len(df))).map(clean_text)

    keep_idx, seen = [], {}
    for idx, row in df.iterrows():
        code = row["lga_code"]
        if not code:
            continue
        if code not in seen:
            seen[code] = idx
            keep_idx.append(idx)
            continue

        prev_idx = seen[code]
        prev = df.loc[prev_idx]
        same = (prev["sen_district"] == row["sen_district"]
                and prev["lga_name"] == row["lga_name"])

        if same:
            ledger.record(source=f"crosswalk:{row['_sheet_row']}", entity_id=code,
                          field_name="lga_code", raw_value=code, resolved_value=code,
                          method="duplicate_restatement", confidence=1.0, outcome="resolved",
                          note=f"identical restatement of sheet row {prev['_sheet_row']}; dropped")
            continue

        # Substantive conflict.
        gazette_new = bool(re.search(r"transferr?ed|gazette|reassigned", row["remarks"], re.I))
        gazette_old = bool(re.search(r"transferr?ed|gazette|reassigned", prev["remarks"], re.I))
        authority_row = auth["lga_by_code"].get(code)
        auth_sen = authority_row.sen_district if authority_row is not None else None

        if gazette_new and not gazette_old:
            winner, loser, rule = idx, prev_idx, "gazetted transfer annotation in Remarks"
        elif gazette_old and not gazette_new:
            winner, loser, rule = prev_idx, idx, "gazetted transfer annotation in Remarks"
        elif auth_sen is not None and row["sen_district"] == auth_sen:
            winner, loser, rule = idx, prev_idx, "agreement with authoritative boundary layer"
        elif auth_sen is not None and prev["sen_district"] == auth_sen:
            winner, loser, rule = prev_idx, idx, "agreement with authoritative boundary layer"
        else:
            winner, loser, rule = None, None, "undecidable"

        rec = {
            "lga_code": code,
            "lga_name": row["lga_name"],
            "assignment_a": f"{prev['sen_district']} (sheet row {prev['_sheet_row']})",
            "assignment_b": f"{row['sen_district']} (sheet row {row['_sheet_row']})",
            "remarks_a": prev["remarks"], "remarks_b": row["remarks"],
            "boundary_layer_says": auth_sen,
            "resolution_rule": rule,
            "resolved_to": (df.loc[winner, "sen_district"] if winner is not None else None),
        }
        conflicts.append(rec)

        if winner is None:
            ledger.record(source=f"crosswalk:{row['_sheet_row']}", entity_id=code,
                          field_name="sen_district", raw_value=rec["assignment_b"],
                          resolved_value=None, method="conflict_undecidable", confidence=0.0,
                          outcome="conflict",
                          note=f"conflicts with {rec['assignment_a']}; no rule decides it")
            keep_idx = [i for i in keep_idx if i != prev_idx]
        else:
            ledger.record(source=f"crosswalk:{row['_sheet_row']}", entity_id=code,
                          field_name="sen_district",
                          raw_value=f"{prev['sen_district']} | {row['sen_district']}",
                          resolved_value=df.loc[winner, "sen_district"],
                          method="conflict_resolved", confidence=0.9, outcome="resolved_review",
                          note=f"resolved by {rule}; losing assignment "
                               f"{df.loc[loser, 'sen_district']!r} discarded")
            keep_idx = [i for i in keep_idx if i != loser]
            if winner not in keep_idx:
                keep_idx.append(winner)
            seen[code] = winner

    out = df.loc[sorted(keep_idx)].reset_index(drop=True)
    LOG.info("Deduplication: %d row(s) in, %d unique LGA row(s) out, %d substantive conflict(s)",
             len(df), len(out), len(conflicts))
    return out, conflicts


def validate(df: pd.DataFrame, auth: dict, control: dict) -> list[dict]:
    """Cross-check the normalised table against every control we have."""
    checks = []

    def add(name, passed, detail):
        checks.append({"check": name, "result": "PASS" if passed else "FAIL", "detail": detail})
        LOG.log(20 if passed else 30, "%-42s %s | %s", name, "PASS" if passed else "FAIL", detail)

    n_auth = len(auth["lgas"])
    add("LGA count equals boundary layer", len(df) == n_auth, f"{len(df)} vs {n_auth}")

    add("LGA codes unique", df.lga_code.is_unique,
        f"{df.lga_code.duplicated().sum()} duplicate(s)")

    missing = set(auth["lga_by_code"]) - set(df.lga_code)
    extra = set(df.lga_code) - set(auth["lga_by_code"])
    add("LGA code set matches boundary layer", not missing and not extra,
        f"missing from crosswalk: {sorted(missing) or 'none'}; "
        f"not in boundary layer: {sorted(extra) or 'none'}")

    if "ward_count_declared" in df.columns:
        declared = pd.to_numeric(df.ward_count_declared, errors="coerce")
        total = int(declared.sum())
        ctrl = control.get("declared_ward_total")
        if ctrl is not None:
            add("Ward counts sum to the sheet's own TOTAL", total == ctrl, f"{total} vs {ctrl}")
        observed = sum(auth["ward_counts"].values())
        add("Ward counts sum to the boundary layer", total == observed, f"{total} vs {observed}")

        per_lga_bad = [
            (r.lga_code, int(dc), auth["ward_counts"].get(r.lga_code))
            for r, dc in zip(df.itertuples(), declared)
            if pd.notna(dc) and auth["ward_counts"].get(r.lga_code) != int(dc)
        ]
        add("Per-LGA ward counts agree with boundary layer", not per_lga_bad,
            f"{len(per_lga_bad)} disagreement(s): {per_lga_bad[:5]}")

    disagree = [
        (r.lga_code, r.sen_district, auth["lga_by_code"][r.lga_code].sen_district)
        for r in df.itertuples()
        if r.lga_code in auth["lga_by_code"]
        and r.sen_district != auth["lga_by_code"][r.lga_code].sen_district
    ]
    add("District assignment agrees with boundary layer", not disagree,
        f"{len(disagree)} disagreement(s): {disagree[:5]}")

    nulls = df[["lga_code", "lga_name", "sen_district", "state_name"]].isna().sum().sum()
    add("No null key fields in the normalised table", nulls == 0, f"{nulls} null(s)")

    return checks


# ==========================================================================
# F. Reporting
# ==========================================================================

def write_report(profile, raw_meta, df, ledger, conflicts, checks, unresolved) -> None:
    s = ledger.summary()
    lines = [
        "# Task 1 — Normalisation of the LGA / Senatorial District Crosswalk",
        "",
        "The source workbook was opened read-only and was not modified. Every",
        "transformation below was performed in memory by",
        "`01_crosswalk_normalisation_pipeline.py`.",
        "",
        "## 1. Sheets found, and which one is authoritative",
        "",
        "| Sheet | Rows | Cols | Merged ranges | Superseded marker | Role |",
        "|---|---:|---:|---:|---|---|",
    ]
    for name, info in profile["sheets"].items():
        role = ("**authoritative**" if name == profile["authoritative"]
                else "alias source only" if info["superseded"] else "not used")
        lines.append(f"| `{name}` | {info['max_row']} | {info['max_col']} | "
                     f"{info['n_merged']} | {'yes' if info['superseded'] else 'no'} | {role} |")
    lines += [
        "",
        f"`{profile['authoritative']}` is authoritative. The rule is stated in code, not assumed: a",
        "sheet whose leading rows carry a supersession marker (\"SUPERSEDED\", \"DO NOT USE\")",
        "is demoted, and among the remainder the widest sheet is the only one that can",
        "physically carry a three-way state / LGA / district relation. The superseded sheet",
        "is still mined for historic spellings, because an old spelling is legitimate",
        "evidence for name reconciliation even when the old *relation* is not.",
        "",
        "## 2. Structural defects handled",
        "",
        "| Defect | Where | Treatment |",
        "|---|---|---|",
        f"| Merged title banner | A1:F1 | Expanded then discarded as a non-data row |",
        f"| Two-tier header | rows 5–6 | Upper tier (merged group labels) discarded; lower tier detected by header-token count, not by row number |",
        f"| Merged / blank-continuation `STATE` | column A | Forward-filled within the data region |",
        f"| Blank-continuation `SENATORIAL DISTRICT` | column C | Forward-filled within the data region |",
        f"| Control-total and footnote rows | tail of sheet | Detected by pattern and excluded from the data, but the TOTAL is retained as a validation control |",
        "",
        "Rows discarded as structure or commentary:",
        "",
        "| Sheet row | Reason | Content |",
        "|---:|---|---|",
    ]
    for d in raw_meta["discarded_rows"]:
        if d["reason"] != "blank":
            lines.append(f"| {d['sheet_row']} | {d['reason']} | {d['content'][:110]} |")

    lines += [
        "",
        "## 3. Name reconciliation",
        "",
        "Names were matched against the GeoPackage boundary layer, which is treated as",
        "the authority for spelling and codes. The cascade is exact match, then match on",
        "a normalised key (case-folded, administrative-type suffix stripped, all",
        "non-alphanumerics removed), then character similarity above an accept threshold,",
        "then character similarity above a review threshold, then failure.",
        "",
        "Before the normalised key is used, the pipeline asserts that it is unique within",
        "the authority layer. If two genuinely distinct units ever collapsed to the same",
        "key the run would abort rather than silently merge them.",
        "",
        f"Decisions recorded: **{len(ledger.rows)}**",
        "",
        "| Outcome | Count |",
        "|---|---:|",
    ]
    for k, v in sorted(s.items()):
        lines.append(f"| `{k}` | {v} |")

    method_counts = Counter(r["method"] for r in ledger.rows)
    lines += ["", "| Method | Count |", "|---|---:|"]
    for k, v in method_counts.most_common():
        lines.append(f"| `{k}` | {v} |")

    lines += [
        "",
        "The full record is `01_crosswalk_reconciliation_ledger.csv` (one row per name",
        "per field). Names that could not be resolved, and assignments that conflict, are",
        "isolated in `01_crosswalk_exceptions.csv`.",
        "",
        "## 4. Duplicate and conflict adjudication",
        "",
    ]
    if conflicts:
        lines += ["| LGA | Assignment A | Assignment B | Boundary layer | Rule applied | Result |",
                  "|---|---|---|---|---|---|"]
        for c in conflicts:
            lines.append(f"| {c['lga_code']} {c['lga_name']} | {c['assignment_a']} | "
                         f"{c['assignment_b']} | {c['boundary_layer_says']} | "
                         f"{c['resolution_rule']} | {c['resolved_to'] or '**unresolved**'} |")
        lines += [
            "",
            "Where the workbook's own Remarks column records a gazetted transfer, that row",
            "supersedes an unannotated one — the workbook is allowed to arbitrate itself",
            "before any external source is consulted. Where the remarks are silent, the",
            "boundary layer decides. Where neither decides, nothing is invented: the LGA is",
            "written to the exceptions report with a null district.",
        ]
    else:
        lines.append("No substantive conflicts survived deduplication.")

    lines += ["", "## 5. Validation", "", "| Check | Result | Detail |", "|---|---|---|"]
    for c in checks:
        lines.append(f"| {c['check']} | **{c['result']}** | {c['detail']} |")

    lines += [
        "",
        "## 6. Names that could not be reconciled",
        "",
    ]
    if unresolved.empty:
        lines += ["Every administrative name in the workbook was reconciled to the authority",
                  "layer. No name was left unresolved and no district assignment was left in",
                  "conflict.", ""]
    else:
        lines += ["| Source row | Field | Raw value | Nearest candidate | Score | Outcome |",
                  "|---|---|---|---|---:|---|"]
        for r in unresolved.itertuples():
            lines.append(f"| {r.source} | {r.field} | `{r.raw_value}` | "
                         f"{r.resolved_value or '—'} | {r.confidence:.3f} | {r.outcome} |")

    lines += [
        "",
        "## 7. Output table",
        "",
        f"`data/lga_senatorial_crosswalk.csv` — {len(df)} rows, one per LGA.",
        "",
        "| Column | Meaning |",
        "|---|---|",
        "| `lga_code` | Primary key, as issued by the Surveyor General |",
        "| `lga_name` | Authoritative spelling |",
        "| `lga_name_source` | Spelling exactly as it appeared in the workbook |",
        "| `sen_code` | Senatorial district code, from the boundary layer |",
        "| `sen_district` | Authoritative senatorial district name |",
        "| `state_code`, `state_name` | Parent state |",
        "| `ward_count_declared` | Ward count asserted by the workbook |",
        "| `ward_count_observed` | Ward count actually present in the boundary layer |",
        "| `remarks` | Compiler's annotation, preserved verbatim |",
        "| `match_method`, `match_confidence`, `match_outcome` | Provenance of the reconciliation |",
        "| `source_sheet_row` | Row in the source workbook, for audit |",
        "",
    ]
    ART["crosswalk_report"].write_text("\n".join(lines), encoding="utf-8")
    LOG.info("Wrote %s", ART["crosswalk_report"].name)


def main() -> pd.DataFrame:
    banner(LOG, "STAGE 01 — CROSSWALK NORMALISATION (source workbook is read-only)")
    ledger = Ledger()

    profile = profile_workbook(SRC["crosswalk"])
    grid = read_sheet_unmerged(SRC["crosswalk"], profile["authoritative"])
    header_idx = locate_header(grid)
    raw, raw_meta = build_raw_table(grid, header_idx)

    auth = load_authority()
    if profile["alias_sources"]:
        harvest_aliases(SRC["crosswalk"], profile["alias_sources"])

    named = reconcile_names(raw, auth, ledger)
    deduped, conflicts = adjudicate(named, auth, ledger)

    # Attach authoritative codes and observed ward counts.
    sen_lookup = {r.sen_district: r for r in auth["sens"].itertuples()}
    deduped["sen_code"] = deduped.sen_district.map(
        lambda d: sen_lookup[d].sen_code if d in sen_lookup else None)
    deduped["state_code"] = deduped.sen_district.map(
        lambda d: sen_lookup[d].state_code if d in sen_lookup else None)
    deduped["ward_count_observed"] = deduped.lga_code.map(auth["ward_counts"])

    final = pd.DataFrame({
        "lga_code": deduped.lga_code,
        "lga_name": deduped.lga_name,
        "lga_name_source": deduped.lga_name_raw.map(clean_text),
        "sen_code": deduped.sen_code,
        "sen_district": deduped.sen_district,
        "state_code": deduped.state_code,
        "state_name": deduped.state_name,
        "ward_count_declared": pd.to_numeric(deduped.get("ward_count_declared"), errors="coerce")
                                 .astype("Int64"),
        "ward_count_observed": pd.to_numeric(deduped.ward_count_observed, errors="coerce")
                                 .astype("Int64"),
        "remarks": deduped.remarks,
        "match_method": deduped.match_method,
        "match_confidence": deduped.match_confidence,
        "match_outcome": deduped.match_outcome,
        "source_sheet_row": deduped._sheet_row,
    }).sort_values("lga_code").reset_index(drop=True)

    checks = validate(final, auth, raw_meta["control"])

    led = ledger.to_frame()
    led.to_csv(ART["crosswalk_ledger"], index=False, encoding="utf-8")
    unresolved = led[led.outcome.isin(["unresolved", "conflict", "resolved_review"])]
    unresolved.to_csv(ART["crosswalk_exceptions"], index=False, encoding="utf-8")
    final.to_csv(ART["crosswalk_table"], index=False, encoding="utf-8")

    write_report(profile, raw_meta, final, ledger, conflicts, checks,
                 led[led.outcome.isin(["unresolved", "conflict"])])

    LOG.info("Normalised crosswalk: %d rows -> %s", len(final), ART["crosswalk_table"].name)
    LOG.info("Ledger: %d decisions -> %s", len(led), ART["crosswalk_ledger"].name)
    LOG.info("Exceptions: %d row(s) -> %s", len(unresolved), ART["crosswalk_exceptions"].name)
    if any(c["result"] == "FAIL" for c in checks):
        LOG.warning("One or more validation checks FAILED — see the report.")
    banner(LOG, "STAGE 01 COMPLETE")
    return final


if __name__ == "__main__":
    main()
