"""
Generate the XLSForm for Form HH/2026 (Bansara State integrated child health and
AMR household survey) and convert it with pyxform.

The form is built from code rather than hand-edited in Excel so that:
  * every constraint has exactly one definition and is exported into the
    constraint register by extract_registers.py - the register cannot drift
    from the form;
  * a mid-round change is a diff in a text file, reviewable in version control,
    rather than an untracked spreadsheet edit (see docs/09_deployment_and_versioning.md);
  * the thresholds live in the CONSTANTS block below, so changing the fieldwork
    window or a plausibility bound is one line, not a hunt through 300 rows.

Run:  python build_xlsform.py

Outputs (into ../form/):
    HH2026_v2-0-0.xlsx      the XLSForm
    HH2026_v2-0-0.xml       the compiled XForm
    conversion_output.txt   pyxform version + full conversion log
"""

from __future__ import annotations

import io
import os
import subprocess
import sys
from contextlib import redirect_stderr, redirect_stdout

try:
    from openpyxl import Workbook
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
FORM_DIR = os.path.normpath(os.path.join(HERE, "..", "form"))

# ===========================================================================
# CONSTANTS - every threshold the questionnaire does not state.
# Each one is justified in docs/02_constraint_register.md by the same key.
# ===========================================================================
FORM_ID = "HH2026"
FORM_TITLE = "Integrated Child Health and AMR Household Survey 2026"
VERSION = "2.0.0"
VERSION_STAMP = "2026060100"  # yyyymmddnn - the value Kobo shows as the version

# ---------------------------------------------------------------------------
# Two deployable variants of ONE instrument.
#
# Both carry both translations and are identical in every question, constraint,
# relevance rule and calculation - `scripts/validate_form.py` runs the full
# structural suite over each of them to prove it. They differ only in which
# language the client opens in, and in form_id.
#
# The separate form_id is the important part. Redeploying a second instrument
# under a live form_id is the Class D breaking change that
# docs/09_deployment_and_versioning.md says never to make mid-round: it would
# either overwrite the deployed Hausa form or merge two instruments into one
# dataset with nothing in the data to tell them apart. Separate ids mean two
# Kobo projects and two datasets, and `form_variant` below is what makes pooling
# them safe afterwards.
# ---------------------------------------------------------------------------
VARIANTS = {
    "ha": {
        "form_id": FORM_ID,
        "title": FORM_TITLE,
        "default_language": "Hausa (ha)",
        "form_variant": "hausa-default",
        "note": "Field deployment. Opens in Hausa; supervisors switch to English.",
    },
    "en": {
        "form_id": f"{FORM_ID}_EN",
        "title": f"{FORM_TITLE} (English)",
        "default_language": "English (en)",
        "form_variant": "english-default",
        "note": "Opens in English; Hausa still available from the client language menu.",
    },
}
DEFAULT_VARIANT = "ha"

FIELDWORK_START = "2026-06-01"  # questionnaire header + ethics approval BSHREC/2026/041
FIELDWORK_END = "2026-06-14"    # operating conditions: "fieldwork runs 14 days"
ETHICS_WINDOW_END = "2026-06-30"  # questionnaire header outer bound - see D-17

HH_SIZE_MIN, HH_SIZE_MAX = 1, 40
STRUCTURE_MIN, STRUCTURE_MAX = 1, 999
SERIAL_MIN, SERIAL_MAX = 1, 999
AGE_MONTHS_MIN, AGE_MONTHS_MAX = 0, 59
AGE_YEARS_MIN, AGE_YEARS_MAX = 5, 97
ELIG_S4_MIN, ELIG_S4_MAX = 9, 59      # Section 4 eligibility, from the questionnaire
ELIG_S5_MIN = 12                      # Section 5 specimen cut, from questionnaire 5.01
N_ELIG_MAX = 20
WEIGHT_MIN, WEIGHT_MAX = 2.0, 30.0
HEIGHT_MIN, HEIGHT_MAX = 45.0, 125.0
POSITION_CUT_MONTHS = 24              # WHO: standing height from 24 months
POSITION_ADJUST_CM = 0.7              # WHO length/height conversion
TEMP_MIN, TEMP_MAX = -5.0, 40.0
TEMP_OK_LO, TEMP_OK_HI = 2.0, 8.0     # cold chain target band
GPS_ACC_MAX_M = 100
GPS_OFFSET_DEG = 0.05                 # ~5.5 km from the settlement centroid
LAT_MIN, LAT_MAX = 10.20, 11.75       # from settlements.csv, widened
LON_MIN, LON_MAX = 6.80, 8.60
DURATION_WARN_MIN = 15                # explanation required below this
DURATION_HARD_MIN = 3                 # cannot be submitted below this
NAME_MAX_CHARS = 15

HA = "label::Hausa (ha)"
EN = "label::English (en)"
HINT_HA = "hint::Hausa (ha)"
HINT_EN = "hint::English (en)"
CM_HA = "constraint_message::Hausa (ha)"
CM_EN = "constraint_message::English (en)"
RM_HA = "required_message::Hausa (ha)"
RM_EN = "required_message::English (en)"

SURVEY_COLS = [
    "type", "name", HA, EN, HINT_HA, HINT_EN,
    "constraint", CM_HA, CM_EN,
    "relevant", "required", RM_HA, RM_EN,
    "calculation", "choice_filter", "appearance", "parameters",
    "default", "read_only", "repeat_count", "trigger",
]
CHOICE_COLS = ["list_name", "name", HA, EN]

survey: list[dict] = []
choices: list[dict] = []

# Registry populated as rows are added; consumed by extract_registers.py.
CONSTRAINT_NOTES: list[dict] = []


def q(type_, name, ha="", en="", **kw):
    """Add a survey row. `note=` records the constraint register entry."""
    note = kw.pop("note", None)
    row = {"type": type_, "name": name, HA: ha, EN: en}
    mapping = {
        "hint_ha": HINT_HA, "hint_en": HINT_EN,
        "cmsg_ha": CM_HA, "cmsg_en": CM_EN,
        "rmsg_ha": RM_HA, "rmsg_en": RM_EN,
    }
    for k, v in kw.items():
        row[mapping.get(k, k)] = v
    survey.append(row)
    if note:
        CONSTRAINT_NOTES.append(
            {
                "id": note[0], "question": note[1], "prevents": note[2], "source": note[3],
                "name": name, "rule": row.get("constraint", "") or row.get("relevant", ""),
            }
        )
    return row


def ch(list_name, name, ha, en):
    choices.append({"list_name": list_name, "name": name, HA: ha, EN: en})


def clist(list_name, items):
    for name, ha, en in items:
        ch(list_name, name, ha, en)


# ===========================================================================
# CHOICE LISTS
# Hausa strings are a WORKING DRAFT prepared by the form developer. They must be
# reviewed and back-translated by a certified Hausa translator before
# deployment - see docs/13_scope_and_exclusions.md, item S-1. The form structure
# makes that a spreadsheet-column task, not a rebuild.
# ===========================================================================
clist("yesno", [("1", "Eh", "Yes"), ("2", "A'a", "No")])
clist("yes_no_dk", [("1", "Eh", "Yes"), ("2", "A'a", "No"), ("8", "Ban sani ba", "Do not know")])
clist("consent", [("1", "An bayar da izini", "Consent given"), ("2", "An ki bayar da izini", "Consent refused")])
clist("sex", [("1", "Namiji", "Male"), ("2", "Mace", "Female")])
clist("relationship", [
    ("1", "Shugaban gida", "Head"),
    ("2", "Mata ko miji", "Spouse"),
    ("3", "Da ko diya", "Son or daughter"),
    ("4", "Uwa ko uba", "Parent"),
    ("5", "Wani dangi", "Other relative"),
    ("6", "Ba dangi ba", "Not related"),
])
clist("age_unit", [
    ("1", "Kasa da shekara 5 - rubuta watanni", "Under 5 years - record completed MONTHS"),
    ("2", "Shekara 5 ko sama - rubuta shekaru", "5 years or older - record completed YEARS"),
])
clist("result_visit", [
    ("1", "An kammala", "Completed"),
    ("2", "An ki", "Refused"),
    ("3", "Babu babban mutum bayan ziyara uku", "No competent adult after three visits"),
    ("4", "Gidan babu kowa ko an rushe", "Dwelling vacant or demolished"),
])
clist("card_seen", [("1", "An ga katin", "Card seen"), ("2", "Ba a ga katin ba", "No card seen")])
clist("record_type", [
    ("1", "Ainihin katin rigakafi", "Original vaccination card"),
    ("2", "Kwafin katin", "Photocopy of the card"),
    ("3", "Rikodin lantarki", "Electronic record"),
])
clist("position", [
    ("1", "A kwance", "Recumbent length"),
    ("2", "A tsaye", "Standing height"),
])
clist("not_measured_reason", [
    ("1", "Mai kula ya ki", "Caregiver refused"),
    ("2", "Yaron ba ya nan", "Child absent"),
    ("3", "Yaron bai natsu ba", "Child would not stay still"),
    ("4", "Na'urar ba ta aiki", "Equipment not working"),
    ("96", "Wani dalili", "Other reason"),
])
clist("photo_status", [
    ("1", "Eh, an dauka", "Yes, photograph taken"),
    ("2", "A'a, kunshin babu", "No, packaging not available"),
    ("3", "Mai kula ya ki", "Caregiver declined"),
])
clist("no_spec_reason", [
    ("1", "Mai kula ya ki", "Caregiver refused"),
    ("2", "Yaron ba ya nan", "Child absent"),
    ("3", "Ba a samu ba", "Unable to produce"),
    ("4", "Kwanon ya lalace", "Container spoiled"),
    ("96", "Wani dalili", "Other"),
])
clist("water", [
    ("1", "Bututun ruwa a cikin gida", "Piped into dwelling"),
    ("2", "Bututun ruwa a harabar gida", "Piped into compound"),
    ("3", "Famfon jama'a", "Public tap or standpipe"),
    ("4", "Rijiyar burtsatse", "Tube well or borehole"),
    ("5", "Rijiya mai kariya", "Protected dug well"),
    ("6", "Rijiya marar kariya", "Unprotected dug well"),
    ("7", "Marmaro mai kariya", "Protected spring"),
    ("8", "Marmaro marar kariya", "Unprotected spring"),
    ("9", "Ruwan sama", "Rainwater"),
    ("10", "Tanka ko keke-ruwa", "Tanker or cart"),
    ("11", "Ruwan saman kasa (kogi, tafki)", "Surface water"),
])
clist("toilet", [
    ("1", "Wanka mai zuwa magudanar ruwa", "Flush to sewer"),
    ("2", "Wanka mai zuwa septic tank", "Flush to septic tank"),
    ("3", "Wanka mai zuwa rami", "Flush to pit latrine"),
    ("4", "Rami mai iska (VIP)", "Ventilated improved pit"),
    ("5", "Rami mai murfi", "Pit latrine with slab"),
    ("6", "Rami marar murfi", "Pit latrine without slab"),
    ("7", "Bandakin takin zamani", "Composting toilet"),
    ("8", "Bokiti", "Bucket"),
    ("9", "Babu bandaki ko daji", "No facility or bush"),
])
clist("handwash", [
    ("1", "An gani, akwai sabulu da ruwa", "Observed, soap and water present"),
    ("2", "An fada kawai, ba a gani ba", "Reported only, not observed"),
    ("3", "Babu", "Not present"),
])
clist("assets", [
    ("A", "Rediyo", "Radio"),
    ("B", "Talabijin", "Television"),
    ("C", "Wayar hannu", "Mobile telephone"),
    ("D", "Keke", "Bicycle"),
    ("E", "Babur", "Motorcycle"),
    ("F", "Mota ko babbar mota", "Car or truck"),
    ("G", "Firij", "Refrigerator"),
    ("H", "Babu ko daya daga cikin wadannan", "None of these"),
])
clist("sup_decision", [
    ("1", "Karba", "Accept"),
    ("2", "Mayar don gyara", "Return for correction"),
    ("3", "Sokewa", "Void"),
])
clist("check_char", [(c, c, c) for c in list("0123456789") + ["X"]])

# ===========================================================================
# XPath fragments
# ===========================================================================
# Weighted sum for the modulus-11 check digit. substr() in ODK is 0-indexed
# with an exclusive end index. Kept identical to scripts/checkdigit.py, and
# test_checkdigit.py T-CD-08 asserts the two agree.
CD_WEIGHTS = (7, 6, 5, 4, 3, 2)
CD_SUM = " + ".join(
    f"number(substr(${{c5_03_serial}}, {i}, {i + 1})) * {w}" for i, w in enumerate(CD_WEIGHTS)
)

LAT = "number(selected-at(${q1_11_gps}, 0))"
LON = "number(selected-at(${q1_11_gps}, 1))"
ACC = "number(selected-at(${q1_11_gps}, 3))"
# Inside a constraint the node must be addressed as '.', so that the value being
# validated is tested rather than the value already committed to the model.
LAT_C, LON_C, ACC_C = (
    "number(selected-at(., 0))", "number(selected-at(., 1))", "number(selected-at(., 3))",
)

# Absolute path to the specimen serial node, written out in full on purpose.
# ${c5_03_serial} referenced from inside its own repeat resolves to '../c5_03_serial',
# which sees only the current repeat instance - a uniqueness test written that way
# always returns 1 and silently checks nothing. The predicate below must be
# evaluated against every instance of the repeat, so the path is absolute.
# validate_form.py asserts this path exists in the compiled XForm, so a change to
# the group nesting breaks the build rather than the check.
SERIAL_ABS = "/data/s3/roster/s5/s5b/c5_03_serial"

IN_HH = "${q1_14_result} = '1' and ${q2_02_consent} = '1'"  # interview actually happened

# ===========================================================================
# SECTION 0 - metadata and operator identification
# ===========================================================================
q("start", "start")
q("end", "end")
q("today", "today")
q("deviceid", "deviceid")
# Audit log: records the order questions were answered and every value change.
# It is the evidence base for the fabrication checks in docs/10_fabrication_detection.md.
# Location tracking inside the audit log was deliberately NOT enabled - see
# docs/11_data_protection.md, DP-7 (continuous staff tracking is disproportionate).
q("audit", "audit", parameters="track-changes=true")

q("calculate", "form_version", calculation=f"'{VERSION}'")
q("calculate", "form_version_stamp", calculation=f"'{VERSION_STAMP}'")
# Which of the two variants produced this record. Overwritten per variant at
# write time. Carried as data so the two datasets can be pooled without
# inferring provenance from the file they arrived in.
q("calculate", "form_variant",
  calculation=f"'{VARIANTS[DEFAULT_VARIANT]['form_variant']}'")

q("begin_group", "s1", "Sashe 1: Bayanin gida", "Section 1: Household identification",
  appearance="field-list" if False else "")

q("note", "s1_note",
  "Cika wannan sashe kafin ka shiga gidan.",
  "Complete this section before entering the dwelling.")

# --- 1.08 / 1.09 moved to the top of Section 1. The paper asks them at 1.08 and
# --- 1.09; on a device the operator must be identified first because the LGA,
# --- ward, supervisor and specimen label block all cascade from it. Flow change
# --- only, no change to any variable. Recorded as CH-1 in the defect register.
q("select_one_from_file staff.csv", "q1_08_enum",
  "1.08 Lambar mai tambaya", "1.08 Enumerator code",
  choice_filter="role='enumerator'", required="yes", appearance="autocomplete",
  hint_ha="Zabi lambarka.", hint_en="Select your own code.")

q("text", "q1_08a_pin", "1.08a Lambar sirri (PIN)", "1.08a Your 4-digit PIN",
  required="yes", appearance="numbers",
  constraint="regex(., '^[0-9]{4}$') and . = pulldata('staff', 'pin', 'name', ${q1_08_enum})",
  cmsg_ha="PIN din bai dace da lambar mai tambaya ba.",
  cmsg_en="That PIN does not match this enumerator code. Only the named enumerator may open a form under their code.",
  note=("C-1.08a", "1.08", "One enumerator completing forms under another's code; unattributable fabrication",
        "Design decision. Uses the pin column already present in staff_roster.csv."))

q("calculate", "calc_team", calculation="pulldata('staff', 'team_code', 'name', ${q1_08_enum})")
q("calculate", "calc_enum_lga", calculation="pulldata('staff', 'lga_code', 'name', ${q1_08_enum})")
q("calculate", "calc_phleb", calculation="pulldata('staff', 'phlebotomy_certified', 'name', ${q1_08_enum})")
q("calculate", "calc_lbl_lo", calculation="pulldata('specimen_alloc', 'range_start', 'name', ${calc_team})")
q("calculate", "calc_lbl_hi", calculation="pulldata('specimen_alloc', 'range_end', 'name', ${calc_team})")

q("note", "q1_09_team_note",
  "1.09 Kungiya: ${calc_team}", "1.09 Team code: ${calc_team}",
  hint_ha="An dauko daga jadawalin ma'aikata.", hint_en="Read from the staff roster. Not keyed.")

q("calculate", "q1_01_state", calculation="'BAN'")

q("select_one_from_file lgas.csv", "q1_02_lga", "1.02 Karamar hukuma", "1.02 Local Government Area",
  required="yes",
  constraint=". = ${calc_enum_lga}",
  cmsg_ha="Ba a sanya ka wannan karamar hukuma ba.",
  cmsg_en="This enumerator is not assigned to that LGA. If the assignment has changed the supervisor must update the staff roster and redeploy the form.",
  note=("C-1.02", "1.02", "Interviews recorded in the wrong LGA; sample frame corruption",
        "staff_roster.csv assigned_lga. Judgement that assignment changes go through the roster file, not free entry."))

q("select_one_from_file wards.csv", "q1_03_ward", "1.03 Unguwa", "1.03 Ward",
  required="yes", choice_filter="lga_code = ${q1_02_lga}", appearance="autocomplete",
  note=("C-1.03", "1.03", "Ward that does not belong to the selected LGA",
        "wards.csv lga_code - referential integrity of the supplied hierarchy."))

q("select_one_from_file settlements.csv", "q1_04_settlement", "1.04 Kauye ko unguwa", "1.04 Settlement",
  required="yes", choice_filter="ward_code = ${q1_03_ward}", appearance="autocomplete",
  hint_ha="Rubuta sunan don nema.", hint_en="Type part of the name to search.",
  note=("C-1.04", "1.04", "Settlement outside the selected ward; free-text settlement names",
        "settlements.csv ward_code. Cascade filter - see docs/05_settlement_serving.md."))

q("calculate", "calc_set_lat", calculation="pulldata('settlements', 'lat', 'name', ${q1_04_settlement})")
q("calculate", "calc_set_lon", calculation="pulldata('settlements', 'lon', 'name', ${q1_04_settlement})")

q("select_one yesno", "q1_05_altname_yn",
  "1.05 Ana kiran wannan wuri da wani suna daban a nan?",
  "1.05 Is the settlement known locally by a different name?", required="yes")
q("text", "q1_05_altname", "1.05a Sunan da ake kira a nan", "1.05a Name used locally",
  relevant="${q1_05_altname_yn} = '1'", required="yes",
  constraint="string-length(.) >= 2 and string-length(.) <= 40",
  cmsg_en="Enter between 2 and 40 characters.",
  cmsg_ha="Rubuta harufa 2 zuwa 40.",
  note=("C-1.05", "1.05", "Single-character placeholders such as 'x' or '-' standing in for a local name that was never asked for",
        "Judgement. A minimum length is the cheapest way to stop a free-text field being dismissed with a keystroke."))

q("integer", "q1_06_structure", "1.06 Lambar da aka rubuta a gidan", "1.06 Structure number painted on the dwelling",
  required="yes", constraint=f". >= {STRUCTURE_MIN} and . <= {STRUCTURE_MAX}",
  cmsg_ha=f"Lambar dole ta kasance tsakanin {STRUCTURE_MIN} da {STRUCTURE_MAX}.",
  cmsg_en=f"Structure number must be between {STRUCTURE_MIN} and {STRUCTURE_MAX}.",
  note=("C-1.06", "1.06", "Keying slips such as 0 or a 4-digit number",
        f"Paper field is 3 boxes, so {STRUCTURE_MAX} is the printed maximum. previous_round_households.csv observed max is 259."))

q("integer", "q1_07_hh_serial", "1.07 Lambar gida a cikin kauyen", "1.07 Household serial number within the settlement",
  required="yes", constraint=f". >= {SERIAL_MIN} and . <= {SERIAL_MAX}",
  cmsg_en=f"Household serial must be between {SERIAL_MIN} and {SERIAL_MAX}.",
  cmsg_ha=f"Lambar gida dole tsakanin {SERIAL_MIN} da {SERIAL_MAX}.",
  note=("C-1.07", "1.07", "Out-of-range serials", "Paper field width (3 boxes)."))

q("date", "q1_10_visit_date", "1.10 Ranar ziyara", "1.10 Date of visit",
  required="yes", default="today()",
  constraint=(f". >= date('{FIELDWORK_START}') and . <= date('{FIELDWORK_END}') and . <= today()"),
  cmsg_ha=f"Ranar ziyara dole ta kasance tsakanin {FIELDWORK_START} da {FIELDWORK_END}, kuma ba ta gaba ba.",
  cmsg_en=(f"Date of visit must fall inside the fieldwork window {FIELDWORK_START} to {FIELDWORK_END} "
           f"and cannot be in the future. If fieldwork has been extended, the form must be reissued."),
  note=("C-1.10", "1.10", "Backdated, forward-dated or out-of-round records; forms completed after the round closed",
        f"Questionnaire header gives 1-30 June 2026; operating conditions give a 14-day round. "
        f"The narrower operational window {FIELDWORK_START}..{FIELDWORK_END} is enforced. See defect D-17."))

q("calculate", "calc_date_not_today", calculation="if(${q1_10_visit_date} = today(), 0, 1)")
q("text", "q1_10a_date_reason",
  "1.10a Me ya sa ranar ziyara ba yau ba ce?", "1.10a Why is the date of visit not today?",
  relevant="${calc_date_not_today} = 1", required="yes",
  constraint="string-length(.) >= 5", cmsg_en="Give a short explanation.",
  cmsg_ha="Bayar da bayani a takaice.",
  note=("C-1.10a", "1.10", "Silent backdating of forms filled in later",
        "Judgement. A visit date other than the device date is legitimate but must be explained."))

q("geopoint", "q1_11_gps", "1.11 GPS a bakin kofar gidan", "1.11 GPS reading at the entrance to the dwelling",
  required="yes",
  constraint=(f"{LAT_C} >= {LAT_MIN} and {LAT_C} <= {LAT_MAX} and "
              f"{LON_C} >= {LON_MIN} and {LON_C} <= {LON_MAX} and {ACC_C} <= {GPS_ACC_MAX_M}"),
  cmsg_ha=f"Wurin ba ya cikin jihar Bansara ko daidaito ya wuce mita {GPS_ACC_MAX_M}. Sake dauka a waje.",
  cmsg_en=(f"The point is outside Bansara State, or accuracy is worse than {GPS_ACC_MAX_M} m. "
           f"Step outside, wait for the accuracy to improve and take the reading again."),
  note=("C-1.11", "1.11", "Readings taken in the wrong state, null islands (0,0), and unusable low-accuracy fixes",
        f"Bounding box widened from settlements.csv (lat {10.37}-{11.57}, lon {6.95}-{8.43}). "
        f"Accuracy ceiling {GPS_ACC_MAX_M} m is judgement."))

q("calculate", "calc_gps_lat", calculation=LAT)
q("calculate", "calc_gps_lon", calculation=LON)
q("calculate", "calc_gps_acc", calculation=ACC)
q("calculate", "calc_flag_gps_far",
  calculation=(f"if(abs({LAT} - number(${{calc_set_lat}})) > {GPS_OFFSET_DEG} or "
               f"abs({LON} - number(${{calc_set_lon}})) > {GPS_OFFSET_DEG}, 1, 0)"))
q("text", "q1_11a_gps_far_reason",
  "1.11a Wurin ya yi nisa da tsakiyar kauyen. Bayyana dalili.",
  "1.11a This point is more than about 5 km from the recorded centre of the selected settlement. Explain.",
  relevant="${calc_flag_gps_far} = 1", required="yes", constraint="string-length(.) >= 5",
  cmsg_en="Give a short explanation.", cmsg_ha="Bayar da bayani a takaice.",
  note=("C-1.11a", "1.04 / 1.11", "The wrong settlement selected from the list; the commonest cause of an unusable record",
        f"Threshold {GPS_OFFSET_DEG} degrees (~5.5 km) against the settlement centroid in settlements.csv. Judgement."))

q("select_one yes_no_dk", "q1_12_prev_round",
  "1.12 An ziyarci wannan gida a zagayen Oktoba 2025?",
  "1.12 Was this household visited during the October 2025 round?", required="yes")

q("select_one yesno", "q1_12a_prev_found",
  "1.12a Gidan yana cikin jerin 2025 na wannan kauyen?",
  "1.12a Is the household in the October 2025 list for this settlement?",
  relevant="${q1_12_prev_round} = '1'", required="yes")

q("select_one_from_file prev_households.csv", "q1_13_prev_hh",
  "1.13 Zabi gidan daga jerin 2025", "1.13 Select the household from the October 2025 register",
  relevant="${q1_12_prev_round} = '1' and ${q1_12a_prev_found} = '1'", required="yes",
  choice_filter="settlement_id = ${q1_04_settlement}", appearance="autocomplete",
  note=("C-1.13", "1.13", "Mis-keyed prior household identifiers, which break the panel linkage entirely",
        "previous_round_households.csv. Selection replaces free text. Rows with consent_to_follow_up = no are "
        "not shipped to the device at all - see docs/11_data_protection.md DP-3."))

q("text", "q1_13b_prev_hh_text",
  "1.13b Rubuta lambar gidan 2025", "1.13b Type the 2025 household identifier",
  relevant="${q1_12_prev_round} = '1' and ${q1_12a_prev_found} = '2'", required="yes",
  constraint="regex(., '^BAN-[0-9]{6}$')",
  cmsg_ha="Tsarin dole ya zama BAN- sai lambobi shida, misali BAN-000123.",
  cmsg_en="The identifier must be BAN- followed by exactly six digits, for example BAN-000123.",
  note=("C-1.13b", "1.13", "Malformed household identifiers when the household is not in the shipped list",
        "Format observed in previous_round_households.csv (BAN-000001)."))

q("calculate", "calc_prev_structure",
  calculation="if(${q1_13_prev_hh} = '', '', pulldata('prev_households', 'structure_number', 'name', ${q1_13_prev_hh}))")
q("calculate", "calc_prev_children",
  calculation="if(${q1_13_prev_hh} = '', '', pulldata('prev_households', 'children_u5_prev', 'name', ${q1_13_prev_hh}))")
q("calculate", "calc_flag_structure_mismatch",
  calculation=("if(${calc_prev_structure} != '' and number(${calc_prev_structure}) != ${q1_06_structure}, 1, 0)"))
q("select_one yesno", "q1_13c_structure_confirm",
  "1.13c Rikodin 2025 ya nuna lamba ${calc_prev_structure}, kai ka rubuta ${q1_06_structure}. Gida daya ne?",
  "1.13c The 2025 record shows structure number ${calc_prev_structure} but you recorded ${q1_06_structure}. Is this the same dwelling?",
  relevant="${calc_flag_structure_mismatch} = 1", required="yes",
  note=("C-1.13c", "1.06 / 1.13", "Linking this interview to the wrong household from the previous round",
        "Cross-check against previous_round_households.structure_number. Warning, not a block: structures are repainted."))

q("select_one result_visit", "q1_14_result", "1.14 Sakamakon ziyara", "1.14 Result of visit", required="yes")
q("note", "q1_14_note",
  "Kada ka cika sashe 2 zuwa 6. Je kai tsaye zuwa Sashe 7.",
  "Do not complete Sections 2 to 6. Go straight to Section 7 and hand the tablet to your supervisor.",
  relevant="${q1_14_result} != '1'")
q("end_group", "s1_end")

# ===========================================================================
# SECTION 2 - consent
# ===========================================================================
q("begin_group", "s2", "Sashe 2: Izini", "Section 2: Consent", relevant="${q1_14_result} = '1'")

q("select_one yesno", "q2_01_consent_read",
  "2.01 An karanta sanarwar izini gaba daya da babbar murya?",
  "2.01 Consent statement read aloud to the respondent in full?", required="yes",
  constraint=". = '1'",
  cmsg_ha="Dole a karanta sanarwar gaba daya kafin a nemi izini. Karanta ta yanzu. Idan ba za a bari ba, koma 1.14 ka rubuta 'An ki'.",
  cmsg_en=("The consent statement must be read in full before consent is sought. Read it now and answer Yes. "
           "If the respondent will not permit it, go back to 1.14 and record the result as Refused."),
  note=("C-2.01", "2.01", "Recording consent that was never validly sought - the paper form attaches no consequence to a No here",
        "Ethics approval BSHREC/2026/041 and defect D-7. Hard block, escalated to the committee for ratification."))

q("select_one yesno", "q2_01a_adult",
  "2.01a Mai amsa yana da shekaru 18 ko sama da haka?",
  "2.01a Is the respondent aged 18 years or older?", required="yes",
  constraint=". = '1'",
  cmsg_ha="Dole babban mutum ya bayar da izini. Idan babu, koma 1.14 ka zabi 'Babu babban mutum'.",
  cmsg_en=("Consent must be given by an adult. If no adult is available, go back to 1.14 and record "
           "'No competent adult after three visits'."),
  note=("C-2.01a", "new", "Consent taken from a minor",
        "ADDED QUESTION, not on the paper form. Ethics requirement; collects a boolean only. Flagged for ratification - defect D-7b."))

q("select_one consent", "q2_02_consent",
  "2.02 Mai amsa ya yarda da hirar gidan?",
  "2.02 Does the respondent consent to the household interview?", required="yes")

q("select_one relationship", "q2_03_relationship",
  "2.03 Alakar mai amsa da shugaban gida",
  "2.03 Relationship of the respondent to the head of household",
  relevant="${q2_02_consent} = '1'", required="yes")
q("end_group", "s2_end")

# ===========================================================================
# SECTION 3 - roster, with the child and specimen modules nested per person
# ===========================================================================
q("begin_group", "s3", "Sashe 3: Jerin mutanen gida", "Section 3: Household roster", relevant=IN_HH)

q("integer", "q3_01_hh_size",
  "3.01 Mutane nawa ne suke zama a wannan gida?",
  "3.01 How many people usually live in this household?", required="yes",
  constraint=f". >= {HH_SIZE_MIN} and . <= {HH_SIZE_MAX}",
  cmsg_ha=f"Adadi dole ya kasance tsakanin {HH_SIZE_MIN} da {HH_SIZE_MAX}.",
  cmsg_en=f"Household size must be between {HH_SIZE_MIN} and {HH_SIZE_MAX}. If the compound is genuinely larger, record it as separate households.",
  note=("C-3.01", "3.01", "Zero-person households and keying slips such as 55 for 5",
        f"Upper bound {HH_SIZE_MAX} is judgement, set well above any plausible single household so it catches keying "
        f"errors rather than large families. The paper field width (2 boxes) would allow 99, which collides with the "
        f"non-response code 99 - see docs/03_coding_and_sentinels.md collision X-6."))

q("note", "s3_roster_intro",
  "Rubuta duk wanda yake zama a nan, farawa daga shugaban gida. Ga kowane yaro kasa da shekara 5, rubuta watanni.",
  "List every usual resident, beginning with the head of household. For a resident under five years old record completed MONTHS; for everyone else record completed YEARS.")

q("begin_repeat", "roster", "Mutum", "Household member")
q("calculate", "r_line", calculation="position(..)")
q("note", "r_line_note", "Layi ${r_line}", "Line ${r_line}")

q("text", "r_name", "(2) Sunan farko da sunan gida", "(2) Initial and family name",
  required="yes",
  hint_ha="Misali: S. Sule. Kada ka rubuta cikakken suna.",
  hint_en="For example 'S. Sule'. Record the first initial and family name only - do not record full given names.",
  constraint=f"string-length(.) >= 2 and string-length(.) <= {NAME_MAX_CHARS} and regex(., \"^[A-Za-z][A-Za-z .'-]*$\")",
  cmsg_ha=f"Harufa 2 zuwa {NAME_MAX_CHARS}, misali S. Sule.",
  cmsg_en=f"Between 2 and {NAME_MAX_CHARS} letters, for example 'S. Sule'. Full given names must not be recorded.",
  note=("C-3.r2", "roster col (2)", "Collection of full personal names that the survey does not need",
        "Data minimisation decision, DP-1. Matches the format already used in previous_round_households.csv."))

q("select_one relationship", "r_relation", "(3) Alaka da shugaban gida", "(3) Relationship to head",
  required="yes",
  constraint="not(${r_line} = 1) or . = '1'",
  cmsg_ha="Layi na daya dole ya zama shugaban gida.",
  cmsg_en="Line 1 must be the head of household - the roster is listed beginning with the head.",
  note=("C-3.r3", "roster col (3)", "Rosters with no head, or a head listed out of order, which breaks relationship coding",
        "Interviewer instruction 'beginning with the head of household'. The paper form prints no coding categories "
        "for this column at all - defect D-12; the list is borrowed from 2.03."))

q("select_one sex", "r_sex", "(4) Jinsi", "(4) Sex", required="yes")

q("select_one age_unit", "r_age_unit", "(5/6) Shekaru ko watanni?", "(5/6) Age recorded in years or months?",
  required="yes",
  hint_ha="Kasa da shekara 5 -> watanni.", hint_en="Under 5 years -> months. 5 years and over -> years.")

q("integer", "r_age_months", "(6) Shekaru cikin watanni", "(6) Age in completed months",
  relevant="${r_age_unit} = '1'", required="yes",
  constraint=f". >= {AGE_MONTHS_MIN} and . <= {AGE_MONTHS_MAX}",
  cmsg_ha=f"Watanni {AGE_MONTHS_MIN} zuwa {AGE_MONTHS_MAX}. Idan yaro ya kai wata 60, zabi shekaru.",
  cmsg_en=f"Months must be {AGE_MONTHS_MIN} to {AGE_MONTHS_MAX}. A child of 60 completed months is 5 years old - record years instead.",
  note=("C-3.r6", "roster col (6)", "Ages in months above 59, which would silently create an ineligible 'eligible' child",
        "Questionnaire: months are recorded for residents under five years."))

q("integer", "r_age_years", "(5) Shekaru", "(5) Age in completed years",
  relevant="${r_age_unit} = '2'", required="yes",
  constraint=f". >= {AGE_YEARS_MIN} and . <= {AGE_YEARS_MAX}",
  cmsg_ha=f"Shekaru {AGE_YEARS_MIN} zuwa {AGE_YEARS_MAX}. Kasa da shekara 5 -> rubuta watanni.",
  cmsg_en=f"Years must be {AGE_YEARS_MIN} to {AGE_YEARS_MAX}. Anyone under 5 years must be recorded in months.",
  note=("C-3.r5", "roster col (5)", "Under-fives recorded in years, which removes them from the eligible pool unnoticed",
        f"Lower bound from the questionnaire's own under-five rule. Upper bound {AGE_YEARS_MAX} is judgement: it stops "
        f"short of 98 and 99 so that a real age can never be confused with the non-response codes - collision X-8."))

q("select_one yesno", "r_age_confirm",
  "Tabbatar: wannan yaro yana da watanni ${r_age_months} daidai?",
  "Confirm: is this child exactly ${r_age_months} completed months old? This decides which questions are asked.",
  relevant=("(${r_age_unit} = '1' and ((${r_age_months} >= 8 and ${r_age_months} <= 13) or ${r_age_months} >= 58)) "
            "or (${r_age_unit} = '2' and ${r_age_years} = 5)"),
  required="yes", constraint=". = '1'",
  cmsg_ha="Sake tambaya sannan ka gyara shekarun.", cmsg_en="Ask again and correct the age before continuing.",
  note=("C-3.rc", "roster cols (5)(6)", "Off-by-one age errors at the two eligibility cuts, which decide whether a child is "
                                        "interviewed at all and whether a specimen is sought",
        "Judgement. Bands 8-13 and 58-59 months straddle the 9-month Section 4 cut, the 12-month Section 5 cut and the "
        "59-month upper cut."))

q("calculate", "r_elig_s4",
  calculation=f"if(${{r_age_unit}} = '1' and ${{r_age_months}} >= {ELIG_S4_MIN} and ${{r_age_months}} <= {ELIG_S4_MAX}, 1, 0)")
q("calculate", "r_elig_s5",
  calculation=f"if(${{r_elig_s4}} = 1 and ${{r_age_months}} >= {ELIG_S5_MIN}, 1, 0)")

# ---------------------------------------------------------------- SECTION 4
q("begin_group", "s4", "Sashe 4: Yaro - ${r_name}", "Section 4: Child module - ${r_name}",
  relevant="${r_elig_s4} = 1")

q("calculate", "c4_01_line", calculation="${r_line}")
q("calculate", "c4_02_name", calculation="${r_name}")
q("calculate", "c4_03_months", calculation="${r_age_months}")
q("calculate", "c4_04_sex", calculation="${r_sex}")
q("note", "c4_intro",
  "4.01-4.04 Yaro: ${r_name}, layi ${r_line}, watanni ${r_age_months}.",
  "4.01-4.04 Child: ${r_name}, roster line ${r_line}, ${r_age_months} months. "
  "These four items are carried from the roster and are not re-keyed.")

q("select_one yesno", "c4_05_measured", "4.05a An auna nauyin yaron?", "4.05a Was the child weighed?",
  required="yes",
  note=("C-4.05a", "4.05", "The 'not measured' sentinel 99 being written into the kilogram field itself",
        "Sentinel separation rule - see docs/03_coding_and_sentinels.md. The paper form carries 99 inside the measurement."))

q("decimal", "c4_05_weight", "4.05 Nauyi (kg)", "4.05 Weight of the child (kg)",
  relevant="${c4_05_measured} = '1'", required="yes",
  constraint=f". >= {WEIGHT_MIN} and . <= {WEIGHT_MAX} and . = round(., 1)",
  cmsg_ha=f"Nauyi {WEIGHT_MIN} zuwa {WEIGHT_MAX} kg, adadi daya bayan digo.",
  cmsg_en=f"Weight must be between {WEIGHT_MIN} and {WEIGHT_MAX} kg and recorded to one decimal place.",
  note=("C-4.05", "4.05", "Decimal-point slips (52.0 for 5.2), scale readings in pounds, and more precision than the scale gives",
        f"Judgement bounds set deliberately wide for children 9-59 months, to catch keying errors rather than clinical "
        f"outliers. Informed by the shape of the WHO weight-for-age standards but not read off them."))

q("select_one not_measured_reason", "c4_05_ns", "4.05b Dalilin rashin awo", "4.05b Reason the child was not weighed",
  relevant="${c4_05_measured} = '2'", required="yes")

q("select_one yesno", "c4_06_measured", "4.06a An auna tsayin yaron?", "4.06a Was the child measured for length or height?",
  required="yes",
  note=("C-4.06a", "4.06", "The 'not measured' sentinel 99 being written into the centimetre field, where 99.0 cm is a "
                           "perfectly ordinary height for a child in this age range",
        "Sentinel separation rule. This is the most damaging collision in the paper instrument - defect D-8."))

q("decimal", "c4_06_height", "4.06 Tsayi (cm)", "4.06 Length or height of the child (cm)",
  relevant="${c4_06_measured} = '1'", required="yes",
  constraint=f". >= {HEIGHT_MIN} and . <= {HEIGHT_MAX} and . = round(., 1)",
  cmsg_ha=f"Tsayi {HEIGHT_MIN} zuwa {HEIGHT_MAX} cm, adadi daya bayan digo.",
  cmsg_en=f"Length or height must be between {HEIGHT_MIN} and {HEIGHT_MAX} cm and recorded to one decimal place.",
  note=("C-4.06", "4.06", "Readings in metres or inches, and transposed digits (150.0 for 105.0)",
        "Judgement bounds spanning a small 9-month-old to a tall 59-month-old, widened at both ends."))

q("select_one not_measured_reason", "c4_06_ns", "4.06b Dalilin rashin awo", "4.06b Reason length or height was not measured",
  relevant="${c4_06_measured} = '2'", required="yes")

q("select_one position", "c4_07_position", "4.07 Yaya aka auna yaron?", "4.07 Position in which the child was measured",
  relevant="${c4_06_measured} = '1'", required="yes")

q("calculate", "calc_pos_expected", calculation=f"if(${{r_age_months}} < {POSITION_CUT_MONTHS}, 1, 2)")
q("select_one yesno", "c4_07_confirm",
  "Yaron yana da watanni ${r_age_months}. Tabbatar da yadda aka auna shi.",
  "This child is ${r_age_months} months old. WHO practice is recumbent length under 24 months and standing height from "
  "24 months. Confirm the position you actually used.",
  relevant=f"${{c4_06_measured}} = '1' and ${{c4_07_position}} != ${{calc_pos_expected}}",
  required="yes", constraint=". = '1'",
  cmsg_ha="Sake duba yadda aka auna.", cmsg_en="Re-check the position before continuing.",
  note=("C-4.07", "4.03 / 4.07", "Length and height being pooled as if interchangeable; they differ systematically by about 0.7 cm",
        f"WHO Child Growth Standards measurement convention: standing height from {POSITION_CUT_MONTHS} months. "
        f"Warning plus a stored adjustment, not a block - a child who cannot stand is measured lying down and that is correct."))
q("calculate", "calc_height_adj_cm",
  calculation=(f"if(${{c4_06_measured}} != '1', '', "
               f"if(${{r_age_months}} < {POSITION_CUT_MONTHS} and ${{c4_07_position}} = '2', ${{c4_06_height}} + {POSITION_ADJUST_CM}, "
               f"if(${{r_age_months}} >= {POSITION_CUT_MONTHS} and ${{c4_07_position}} = '1', ${{c4_06_height}} - {POSITION_ADJUST_CM}, "
               f"${{c4_06_height}})))"))

q("select_one card_seen", "c4_08_card",
  "4.08 Zan iya ganin katin rigakafin yaron?", "4.08 May I see the child's vaccination card or health record?",
  required="yes",
  hint_ha="Ka gani da idonka.", hint_en="Record what you personally saw.")
q("select_one record_type", "c4_08a_record_type",
  "4.08a Wanne irin rikodi ka gani?", "4.08a Which kind of record did you see?",
  relevant="${c4_08_card} = '1'", required="yes",
  note=("C-4.08a", "4.08", "The question text asks the interviewer to distinguish a card, a card copy and an electronic "
                           "record, but the printed codes can record only 'seen' or 'not seen' - the distinction is lost",
        "ADDED sub-question, defect D-11. The approved 1/2 coding at 4.08 is untouched."))

q("select_one yesno", "c4_09_measles_card",
  "4.09 Daga katin: an rubuta allurar kyanda?", "4.09 Copy from the card: is a measles dose recorded?",
  relevant="${c4_08_card} = '1'", required="yes")
q("select_one yes_no_dk", "c4_10_measles_report",
  "4.10 An taba yi wa yaron allurar kyanda?", "4.10 Has this child ever received a measles vaccination?",
  relevant="${c4_08_card} = '2'", required="yes")
q("calculate", "calc_measles_any",
  calculation="if(${c4_08_card} = '1', ${c4_09_measles_card}, ${c4_10_measles_report})")
q("calculate", "calc_measles_source",
  calculation="if(${c4_08_card} = '1', 'card', if(${c4_08_card} = '2', 'recall', ''))")

q("select_one yes_no_dk", "c4_11_diarrhoea",
  "4.11 Yaron ya yi gudawa cikin kwanaki 14 da suka wuce?",
  "4.11 Has this child had diarrhoea in the past 14 days?", required="yes")

q("select_one yes_no_dk", "c4_12_antibiotic",
  "4.12 Yaron ya sha maganin kashe kwayoyin cuta cikin kwanaki 30 da suka wuce?",
  "4.12 Has this child taken any antibiotic medicine in the past 30 days?", required="yes")

q("select_one_from_file medicines.csv", "c4_13_medicine",
  "4.13 Wanne magani ne?", "4.13 Which antibiotic was taken?",
  relevant="${c4_12_antibiotic} = '1'", required="yes", appearance="autocomplete",
  hint_ha="Idan fiye da daya, rubuta na baya-bayan nan.",
  hint_en="Where more than one was taken, record the most recent.",
  note=("C-4.13", "4.13", "Free-text drug names that cannot be coded to the AWaRe classification",
        "PLACEHOLDER medicine list - the questionnaire refers to a medicine list that does not exist anywhere in the "
        "data pack. Blocking defect D-15. Codes avoid 96/98/99."))
q("text", "c4_14_medicine_other", "4.14 Rubuta sunan maganin", "4.14 Write the name of the medicine as reported",
  relevant="${c4_13_medicine} = '96'", required="yes",
  constraint="string-length(.) >= 2 and string-length(.) <= 60",
  cmsg_en="Between 2 and 60 characters.", cmsg_ha="Harufa 2 zuwa 60.",
  note=("C-4.14", "4.14", "An 'Other' antibiotic with no name written against it - the response is then uncodeable and the "
                          "record is lost to the AMR analysis entirely",
        "Questionnaire note: 'that category is Other and the response must be written in full on the line provided'. "
        "The paper form states the requirement but cannot enforce it."))
q("select_one yes_no_dk", "c4_15_no_prescription",
  "4.15 An samu maganin ba tare da takardar likita ba?",
  "4.15 Was the medicine obtained without a prescription from a health worker?",
  relevant="${c4_12_antibiotic} = '1'", required="yes")
q("select_one photo_status", "c4_16_photo_status",
  "4.16 An dauki hoton kunshin maganin?", "4.16 Was a photograph of the medicine packaging taken?",
  relevant="${c4_12_antibiotic} = '1'", required="yes")
q("image", "c4_16a_photo", "4.16a Hoton kunshin maganin", "4.16a Photograph of the medicine packaging",
  relevant="${c4_16_photo_status} = '1'", required="yes", parameters="max-pixels=1024",
  hint_ha="Dauki hoton kunshin kadai. Kada ka sanya fuska ko mutum a hoton.",
  hint_en="Photograph the packaging only. No faces, hands or people in the frame.",
  note=("C-4.16a", "4.16", "A 'photograph taken' answer with no photograph attached to the record, and incidental capture "
                           "of identifiable people",
        "The paper form can only record that a photograph exists elsewhere. Resolution size capped at 1024 px for the "
        "9-day offline storage budget - see docs/09_deployment_and_versioning.md."))
q("end_group", "s4_end")

# ---------------------------------------------------------------- SECTION 5
q("begin_group", "s5", "Sashe 5: Tattara samfuri", "Section 5: Specimen collection",
  relevant="${r_elig_s4} = 1")

q("note", "c5_01_note",
  "5.01 Yaron yana da watanni ${r_age_months}.",
  "5.01 This child is ${r_age_months} completed months old. "
  "A specimen is sought from children aged 12 completed months or older.")
q("calculate", "c5_01_age12", calculation="${r_elig_s5}")
q("note", "c5_01_skip",
  "Yaron bai kai wata 12 ba. Ba a neman samfuri. Ci gaba zuwa yaro na gaba.",
  "This child is under 12 completed months. No specimen is sought. Continue to the next household member.",
  relevant="${r_elig_s5} = 0")

q("begin_group", "s5b", "", "", relevant="${r_elig_s5} = 1")

q("select_one yesno", "c5_02_obtained",
  "5.02 An samu samfurin bayan gida daga wannan yaro?",
  "5.02 Was a stool specimen obtained from this child?", required="yes",
  note=("C-5.02", "5.02", "The paper form prints no skip instruction here at all, so 5.03-5.05 (label, time, temperature) "
                          "and 5.06-5.07 (reason none obtained) are both left open whichever way it is answered",
        "Defect D-3, resolved in the form: Yes opens 5.03-5.05, No opens 5.06-5.07, and neither branch can be left blank."))

q("text", "c5_03_serial", "5.03 Lambar samfurin (lambobi shida)", "5.03 Specimen label number (six digits)",
  relevant="${c5_02_obtained} = '1'", required="yes", appearance="numbers",
  hint_ha="Manna lambar a kwanon, sannan ka rubuta lambobin shida.",
  hint_en="Affix the label to the container first, then key the six digits printed on it.",
  constraint=("regex(., '^[0-9]{6}$') "
              "and number(.) >= number(${calc_lbl_lo}) and number(.) <= number(${calc_lbl_hi}) "
              f"and count({SERIAL_ABS}[. = current()/.]) = 1"),
  cmsg_ha="Lambar dole ta kasance cikin lambobin kungiyar ${calc_team} (${calc_lbl_lo} zuwa ${calc_lbl_hi}) kuma ba a sha biyu ba.",
  cmsg_en=("The serial must be six digits, must fall inside the block issued to team ${calc_team} "
           "(${calc_lbl_lo} to ${calc_lbl_hi}), and must not already have been used for another child in this household."),
  note=("C-5.03a", "5.03", "Labels from another team's block, six-digit typos outside the issued range, and the same label "
                           "recorded twice within one household",
        "specimen_label_allocation.csv range_start/range_end for the enumerator's team. Within-submission uniqueness is "
        "the only duplicate check a self-contained form can make - see docs/06_specimen_labels.md."))

q("calculate", "calc_cd_sum", calculation=CD_SUM)
q("calculate", "calc_cd_expected",
  calculation="if(string-length(${c5_03_serial}) != 6, '', if(${calc_cd_sum} mod 11 = 10, 'X', string(${calc_cd_sum} mod 11)))")

q("select_one check_char", "c5_03_check", "5.03a Lambar tantancewa (bayan layin)", "5.03a Check digit (after the dash)",
  relevant="${c5_02_obtained} = '1'", required="yes", appearance="horizontal-compact",
  hint_ha="Zabi harafi ko lambar da ke bayan layin a lambar da aka buga.",
  hint_en="Tap the character printed after the dash on the label.",
  constraint=". = ${calc_cd_expected}",
  cmsg_ha="Lambar tantancewa ba ta dace ba. Sake karanta lambar da ke kan kwanon. Kada ka canza lamba don ta wuce.",
  cmsg_en=("The check digit does not match the serial. Read the label on the container again and re-key the six digits. "
           "Do not change the check digit to make it pass."),
  note=("C-5.03b", "5.03", "Any single mis-keyed digit and ANY transposition of two digits in the serial - a specimen that "
                           "cannot be matched to a child record is discarded and the child must be revisited",
        "Check digit scheme stated in specimen_label_allocation.csv: modulus 11, weights 2 to 7 right to left, remainder "
        "10 recorded as X. Demonstrated over 292,960 transposition cases in scripts/test_checkdigit.py."))

q("calculate", "calc_label_full",
  calculation="if(${c5_03_serial} = '', '', concat('BSN', ${c5_03_serial}, '-', ${c5_03_check}))")

q("time", "c5_04_time", "5.04 Lokacin sanya samfurin a cikin sanyi", "5.04 Time the specimen was placed in the cold box",
  relevant="${c5_02_obtained} = '1'", required="yes",
  note=("C-5.04", "5.04", "Missing cold chain timing, which makes a specimen unusable for the laboratory",
        "Questionnaire 5.04. Required whenever a specimen was obtained."))

q("decimal", "c5_05_temp", "5.05 Zafin akwatin sanyi (C)", "5.05 Temperature shown on the cold box thermometer (C)",
  relevant="${c5_02_obtained} = '1'", required="yes",
  constraint=f". >= {TEMP_MIN} and . <= {TEMP_MAX} and . = round(., 1)",
  cmsg_ha=f"Zafi tsakanin {TEMP_MIN} da {TEMP_MAX} digiri, adadi daya bayan digo.",
  cmsg_en=f"Temperature must be between {TEMP_MIN} and {TEMP_MAX} C, to one decimal place.",
  note=("C-5.05", "5.05", "A cold chain breach that cannot be written down at all: the printed field is one digit and a "
                          "decimal, so no reading of 10 C or above can be recorded",
        f"Defect D-10. Range widened to {TEMP_MIN}..{TEMP_MAX} C, which is the range a cold box thermometer can show."))

q("select_one yesno", "c5_05_confirm",
  f"Zafin ya fita daga {TEMP_OK_LO} zuwa {TEMP_OK_HI} digiri. Tabbatar da karatun sannan ka gaya wa shugaba yau.",
  f"The cold box is outside the {TEMP_OK_LO} to {TEMP_OK_HI} C target band. Confirm the reading and tell your supervisor today.",
  relevant=f"${{c5_02_obtained}} = '1' and (${{c5_05_temp}} < {TEMP_OK_LO} or ${{c5_05_temp}} > {TEMP_OK_HI})",
  required="yes", constraint=". = '1'",
  cmsg_en="Re-read the thermometer before continuing.", cmsg_ha="Sake duba ma'aunin zafi.",
  note=("C-5.05a", "5.05", "Silent cold chain failure - specimens arriving warm with nothing in the record to explain it",
        f"Cold chain target band {TEMP_OK_LO}-{TEMP_OK_HI} C for specimen transport. Warning only, so the true reading is "
        f"still recorded."))

q("select_one no_spec_reason", "c5_06_reason", "5.06 Dalilin rashin samun samfuri", "5.06 Reason no specimen was obtained",
  relevant="${c5_02_obtained} = '2'", required="yes")
q("text", "c5_07_other", "5.07 Bayyana dalilin", "5.07 If code 96, specify",
  relevant="${c5_06_reason} = '96'", required="yes",
  constraint="string-length(.) >= 3 and string-length(.) <= 60",
  cmsg_en="Between 3 and 60 characters.", cmsg_ha="Harufa 3 zuwa 60.",
  note=("C-5.07", "5.07", "An unexplained 'Other' reason for a missing specimen, which is the one field that would tell the "
                          "laboratory why collection is failing",
        "Questionnaire note on code 96. Enforced rather than requested."))
q("end_group", "s5b_end")
q("end_group", "s5_end")
q("end_repeat", "roster_end")

# --------------------------------------------------- roster reconciliation
q("calculate", "calc_n_roster", calculation="count(${r_line})")
q("calculate", "calc_n_head", calculation="count(${r_relation}[. = '1'])")
q("calculate", "calc_n_elig", calculation="count(${r_elig_s4}[. = 1])")
q("calculate", "calc_n_elig_s5", calculation="count(${r_elig_s5}[. = 1])")
q("calculate", "calc_n_modules", calculation="count(${c4_08_card}[. != ''])")
q("calculate", "calc_n_specimens", calculation="count(${c5_03_serial}[. != ''])")
q("calculate", "calc_n_cards_seen", calculation="count(${c4_08_card}[. = '1'])")
q("calculate", "calc_n_weighed", calculation="count(${c4_05_weight}[. != ''])")

q("integer", "q3_01_gate",
  "Ka ce mutane ${q3_01_hh_size} ne, amma ka rubuta ${calc_n_roster}. Gyara jerin ko adadin, sannan ka rubuta adadin da ya dace a nan.",
  "You said ${q3_01_hh_size} people live here but you have listed ${calc_n_roster}. Go back and add the missing people, "
  "or correct 3.01. When the roster is right, type the number of people listed here.",
  relevant="${calc_n_roster} != ${q3_01_hh_size}", required="yes",
  constraint=". = ${calc_n_roster} and ${calc_n_roster} = ${q3_01_hh_size}",
  cmsg_ha="Adadin da ka rubuta a 3.01 dole ya yi daidai da adadin mutanen da ke jerin.",
  cmsg_en="3.01 and the roster must agree before you can continue.",
  note=("C-3.01g", "3.01 vs roster", "The single most common household survey error: a stated household size that does not "
                                     "match the people actually listed. On paper a clerk may or may not notice",
        "Required by the brief. Hard gate: the question disappears only when the two agree."))

q("integer", "q3_head_gate",
  "Jerin dole ya kasance da shugaban gida daya tak. Yanzu akwai ${calc_n_head}. Gyara jerin sannan ka rubuta 1.",
  "The roster must contain exactly one head of household. It currently has ${calc_n_head}. Correct the roster, then type 1 here.",
  relevant="${calc_n_head} != 1", required="yes",
  constraint=". = 1 and ${calc_n_head} = 1",
  cmsg_en="Correct the roster so that exactly one member is coded as head.",
  cmsg_ha="Gyara jerin har sai shugaba daya tak ya rage.",
  note=("C-3.head", "roster col (3)", "Rosters with no head or two heads, which make relationship variables unusable",
        "Judgement, following the interviewer instruction to begin with the head."))

q("integer", "q3_02_stated",
  "3.02 Yara nawa ne a wannan gida masu watanni 9 zuwa 59?",
  "3.02 How many children in this household are aged 9 to 59 completed months?",
  required="yes", constraint=f". >= 0 and . <= {N_ELIG_MAX}",
  cmsg_en=f"Enter a number between 0 and {N_ELIG_MAX}.", cmsg_ha=f"Rubuta adadi tsakanin 0 da {N_ELIG_MAX}.",
  note=("C-3.02", "3.02", "Implausible eligible-child counts",
        "The paper form asks the interviewer to read this off column (7), which is marked office use and must be left "
        "blank in the field - defect D-1. Here it is asked independently and then reconciled."))

q("integer", "q3_02_gate",
  "Daga jerin, yara ${calc_n_elig} ne ke da watanni 9-59, amma ka ce ${q3_02_stated}. Gyara sannan ka rubuta adadin da ya dace.",
  "From the roster, ${calc_n_elig} children are aged 9 to 59 completed months, but you answered ${q3_02_stated}. "
  "Correct the ages in the roster or correct 3.02, then type the right number here.",
  relevant="${q3_02_stated} != ${calc_n_elig}", required="yes",
  constraint=". = ${calc_n_elig} and ${calc_n_elig} = ${q3_02_stated}",
  cmsg_ha="Adadin 3.02 dole ya yi daidai da jerin.",
  cmsg_en="3.02 and the roster must agree before you can continue.",
  note=("C-3.02g", "3.02 vs Section 4", "A stated number of eligible children that differs from the number of child modules "
                                        "completed - the error that leaves a child uninterviewed and a specimen unmatchable",
        "Required by the brief. The child module is nested inside the roster and opens automatically for every eligible "
        "child, so modules completed and eligible children cannot diverge once 3.02 agrees."))

q("note", "s3_no_children",
  "Babu yaro mai watanni 9-59 a wannan gida. Ci gaba zuwa Sashe 6.",
  "No child aged 9 to 59 completed months lives here. Continue to Section 6.",
  relevant="${calc_n_elig} = 0")
q("end_group", "s3_end")

# ===========================================================================
# SECTION 6 - household environment
# ===========================================================================
q("begin_group", "s6", "Sashe 6: Yanayin gida", "Section 6: Household environment", relevant=IN_HH)

q("select_one water", "q6_01_water",
  "6.01 Daga ina kuke samun ruwan sha?",
  "6.01 What is the main source of drinking water for members of this household?", required="yes",
  note=("C-6.01", "6.01", "A 'do not know' coded as 8 being read as 'unprotected spring', and 9 as 'rainwater' - the "
                          "questionnaire's global sentinel codes collide with two substantive categories here",
        "Collision X-1. Resolved by not offering a do-not-know option on an observable characteristic; if one is ever "
        "required it must use a reserved code, never 8 or 9."))

q("select_one toilet", "q6_02_toilet",
  "6.02 Wanne irin bandaki kuke amfani da shi?",
  "6.02 What kind of toilet facility do members of this household usually use?", required="yes",
  note=("C-6.02", "6.02", "Same collision as 6.01: 8 is 'bucket' and 9 is 'no facility or bush'",
        "Collision X-2."))

q("select_one yesno", "q6_03_animals",
  "6.03 Kuna kiwon kaji ko dabbobi a cikin harabar gidan?",
  "6.03 Does this household keep poultry or livestock inside the compound?", required="yes")
q("select_one yes_no_dk", "q6_04_animal_abx",
  "6.04 An ba wa dabbobin maganin kashe kwayoyin cuta cikin watanni 12 da suka wuce?",
  "6.04 Have any antibiotic medicines been given to these animals in the past 12 months?",
  relevant="${q6_03_animals} = '1'", required="yes")

q("select_one handwash", "q6_05_handwash",
  "6.05 Duba: akwai wurin wanke hannu da sabulu da ruwa?",
  "6.05 Observe: is there a handwashing station with both soap and water available?", required="yes")

q("select_one yes_no_dk", "q6_06_hh_diarrhoea",
  "6.06 Wani a cikin gidan ya yi gudawa cikin makonni biyu da suka wuce?",
  "6.06 Has any member of this household had diarrhoea in the past two weeks?", required="yes")

q("calculate", "calc_child_diarrhoea", calculation="count(${c4_11_diarrhoea}[. = '1'])")
q("select_one yesno", "q6_06_confirm",
  "Ka rubuta cewa yaro ${calc_child_diarrhoea} ya yi gudawa a 4.11, amma ka ce babu a 6.06. Tabbatar.",
  "You recorded diarrhoea for ${calc_child_diarrhoea} child(ren) at 4.11 but answered No at 6.06. "
  "A child is a member of the household. Confirm you have re-asked the question.",
  relevant="${calc_child_diarrhoea} > 0 and ${q6_06_hh_diarrhoea} = '2'", required="yes", constraint=". = '1'",
  cmsg_en="Re-ask 6.06 and correct 4.11 or 6.06 before continuing.",
  cmsg_ha="Sake tambaya 6.06 sannan ka gyara.",
  note=("C-6.06", "4.11 vs 6.06", "A logically impossible household in which a child has diarrhoea but no household member does",
        "Cross-question consistency. Judgement; a warning rather than a block because 6.06 is asked of the respondent, "
        "who may not know about a child's illness."))

q("select_multiple assets", "q6_07_assets",
  "6.07 Wanne daga cikin wadannan wannan gida yake da shi?",
  "6.07 Which of the following does this household own? Record all that apply.", required="yes",
  constraint="not(selected(., 'H') and count-selected(.) > 1)",
  cmsg_ha="'Babu ko daya' ba za a zaba tare da wani abu ba.",
  cmsg_en="'None of these' cannot be selected together with any other item.",
  note=("C-6.07", "6.07", "Records that simultaneously own a radio and own nothing",
        "Judgement. 'None of these' is a substantive category on paper with nothing preventing it being ticked alongside others."))
q("end_group", "s6_end")

# ===========================================================================
# SECTION 7 - close-out and supervisor review
# ===========================================================================
q("begin_group", "s7", "Sashe 7: Kammalawa da nazarin shugaba", "Section 7: Close-out and supervisor review")

q("time", "q7_01_end_time", "7.01 Lokacin da hirar ta kare", "7.01 Time the interview ended",
  required="yes", relevant="${q1_14_result} = '1'")

q("calculate", "calc_duration_min",
  calculation=("if(${q7_01_end_time} = '', '', "
               "int((decimal-date-time(now()) - decimal-date-time(${start})) * 1440))"))

q("text", "q7_01a_short_reason",
  "7.01a Hirar ta dauki mintuna ${calc_duration_min} kawai. Bayyana dalili.",
  "7.01a This interview took only ${calc_duration_min} minutes. Explain why it was shorter than usual.",
  relevant=f"{IN_HH} and ${{calc_duration_min}} != '' and ${{calc_duration_min}} < {DURATION_WARN_MIN}",
  required="yes", constraint="string-length(.) >= 10",
  cmsg_en="Give a real explanation of at least 10 characters.", cmsg_ha="Bayar da bayani na gaskiya.",
  note=("C-7.01a", "7.01", "The pattern described in the operating conditions: 94 interviews averaging 4 minutes, "
                           "discovered only after fieldwork closed",
        f"Judgement, calibrated on the reported incident. {DURATION_WARN_MIN} minutes is roughly the fastest a genuine "
        f"interview with one child module and a specimen can be completed."))

q("integer", "q7_01b_hard_gate",
  "Hirar ta dauki mintuna ${calc_duration_min}. Ba za a iya aika wannan fom ba. Kira shugabanka yanzu.",
  "This interview took ${calc_duration_min} minutes, which is below the minimum for a completed interview. "
  "This form cannot be finalised. Call your supervisor now.",
  relevant=f"{IN_HH} and ${{calc_duration_min}} != '' and ${{calc_duration_min}} < {DURATION_HARD_MIN}",
  required="yes", constraint=f"${{calc_duration_min}} >= {DURATION_HARD_MIN}",
  cmsg_en="A completed household interview cannot take less than three minutes.",
  cmsg_ha="Hira da aka kammala ba za ta gaza mintuna uku ba.",
  note=("C-7.01b", "7.01", "Bulk fabrication of completed interviews in minutes",
        f"Judgement. Hard floor {DURATION_HARD_MIN} minutes applies only to interviews recorded as completed with consent "
        f"given, so a genuine refusal at 1.14 is never blocked."))

q("text", "q7_02_observations", "7.02 Duk wani abin lura", "7.02 Any observation that may help the office interpret this form",
  appearance="multiline")

q("select_one yesno", "q7_03_attest",
  "7.03 Na tabbatar da cewa ni da kaina na yi wannan hira, kuma amsoshin sun fito daga mai amsa.",
  "7.03 I confirm that I personally conducted this interview and that the answers are as the respondent gave them.",
  required="yes", constraint=". = '1'",
  cmsg_ha="Dole ka tabbatar kafin a aika.", cmsg_en="You must attest before the form can be finalised.",
  note=("C-7.03a", "7.03", "An unattested form. On paper an unsigned form is visible in the pile; in a digital submission "
                           "nothing marks the absence unless it is required",
        "Questionnaire 7.03 requires a signature. Rendered as an explicit attestation plus a PIN - see DP-5."))
q("text", "q7_03_pin", "7.03a Sake shigar da PIN dinka", "7.03a Re-enter your PIN to sign",
  required="yes", appearance="numbers",
  constraint=". = pulldata('staff', 'pin', 'name', ${q1_08_enum})",
  cmsg_ha="PIN bai dace ba.", cmsg_en="PIN does not match the enumerator recorded at 1.08.",
  note=("C-7.03", "7.03", "An unsigned form, and a form signed by someone other than the named enumerator",
        "Replaces the ink signature. A signature image was rejected on data protection grounds - DP-5."))
q("calculate", "q7_03_date", calculation="today()")

q("select_one yesno", "q7_sup_present",
  "Shugaba yana nan yanzu don duba wannan fom?", "Is the supervisor here now to review this form?",
  required="yes")
q("note", "q7_sup_absent_note",
  "Za a duba wannan fom a ofishi. Kada ka goge shi daga tablet din.",
  "This form will be reviewed at the hub. Do not delete it from the tablet until it has been sent and accepted.",
  relevant="${q7_sup_present} = '2'")

q("begin_group", "s7sup", "Nazarin shugaba", "Supervisor review", relevant="${q7_sup_present} = '1'")
q("select_one_from_file staff.csv", "q7_04_sup", "7.04 Lambar shugaba", "7.04 Supervisor code",
  required="yes", choice_filter="role='supervisor' and team_code=${calc_team}", appearance="autocomplete",
  note=("C-7.04", "7.04", "An enumerator signing off their own work: the paper form prints the same ENU prefix for the "
                          "supervisor code and nothing distinguishes the two",
        "staff_roster.csv role and team_code. The list is filtered to the supervisor of this enumerator's own team."))
q("text", "q7_04a_sup_pin", "7.04a PIN na shugaba", "7.04a Supervisor PIN",
  required="yes", appearance="numbers",
  constraint=". = pulldata('staff', 'pin', 'name', ${q7_04_sup})",
  cmsg_ha="PIN na shugaba bai dace ba. Shugaba da kansa ne zai shigar da nasa PIN.",
  cmsg_en="Supervisor PIN incorrect. The supervisor must enter their own PIN.",
  note=("C-7.04a", "7.04", "Self-approval of fabricated work",
        "Design decision, using the pin column in staff_roster.csv."))

q("note", "q7_qa_summary",
  "Takaitawa ga shugaba.",
  "SUPERVISOR REVIEW SUMMARY\n"
  "Duration: ${calc_duration_min} min | Household size: ${calc_n_roster} | Eligible children: ${calc_n_elig}\n"
  "Child modules completed: ${calc_n_modules} | Vaccination cards seen: ${calc_n_cards_seen} | Children weighed: ${calc_n_weighed}\n"
  "Specimens collected: ${calc_n_specimens} of ${calc_n_elig_s5} children aged 12 months and over\n"
  "GPS more than 5 km from settlement centre: ${calc_flag_gps_far} | GPS accuracy: ${calc_gps_acc} m")

q("select_one sup_decision", "q7_05_decision", "7.05 Shawarar shugaba kan wannan fom",
  "7.05 Supervisor decision on this form", required="yes")
q("text", "q7_05a_reason", "7.05a Dalili", "7.05a Reason for returning or voiding this form",
  relevant="${q7_05_decision} != '1'", required="yes", constraint="string-length(.) >= 10",
  cmsg_en="Give a reason of at least 10 characters.", cmsg_ha="Bayar da dalili.",
  note=("C-7.05a", "7.05", "Forms returned or voided with no recorded reason, which cannot be actioned or audited",
        "Judgement."))
q("select_one yesno", "q7_06_attest",
  "7.06 Na duba wannan fom da kaina.", "7.06 I have reviewed this form myself.",
  required="yes", constraint=". = '1'",
  cmsg_en="The supervisor must attest before the form can be finalised.", cmsg_ha="Dole shugaba ya tabbatar.",
  note=("C-7.06", "7.06", "A supervisor decision recorded at 7.05 with no accompanying attestation that the form was "
                          "actually looked at",
        "Questionnaire 7.06. Paired with the supervisor PIN at 7.04a so the attestation is attributable."))
q("calculate", "q7_06_date", calculation="today()")
q("end_group", "s7sup_end")
q("end_group", "s7_end")

# ===========================================================================
# EXPORT FLAGS - computed once, carried on every submission, consumed by
# scripts/daily_qa_checks.py. Nothing here is asked of anyone.
#
# These are deliberately NOT wrapped in a group. A group containing only
# calculates produces a <group> element with no children in the compiled body,
# and ODK Validate rejects that outright ("Group has no children!"), so the form
# will not deploy. Calculates are bind-only and need no body container; the
# grouping was cosmetic. validate_form.py V-17 now fails the build on any empty
# body group, so this cannot recur.
# ===========================================================================
q("calculate", "calc_hh_key",
  calculation="concat(${q1_04_settlement}, '-', ${q1_06_structure}, '-', ${q1_07_hh_serial})")
q("calculate", "flag_short_interview",
  calculation=f"if(${{calc_duration_min}} != '' and ${{calc_duration_min}} < {DURATION_WARN_MIN}, 1, 0)")
q("calculate", "flag_no_cards_seen",
  calculation="if(${calc_n_modules} > 0 and ${calc_n_cards_seen} = 0, 1, 0)")
q("calculate", "flag_roster_corrected",
  calculation="if(${q3_01_gate} != '' or ${q3_02_gate} != '', 1, 0)")
q("calculate", "flag_specimen_shortfall",
  calculation="if(${calc_n_elig_s5} > 0 and ${calc_n_specimens} = 0, 1, 0)")
q("calculate", "flag_cold_chain",
  calculation=f"if(count(${{c5_05_temp}}[. > {TEMP_OK_HI}]) + count(${{c5_05_temp}}[. < {TEMP_OK_LO}]) > 0, 1, 0)")


# ===========================================================================
# WRITE + CONVERT
# ===========================================================================
def write_xlsx(path: str, variant: str = DEFAULT_VARIANT) -> None:
    v = VARIANTS[variant]
    wb = Workbook()
    ws = wb.active
    ws.title = "survey"
    ws.append(SURVEY_COLS)
    for r in survey:
        row = dict(r)
        if row["name"] == "form_variant":
            row["calculation"] = f"'{v['form_variant']}'"
        ws.append([row.get(c, "") for c in SURVEY_COLS])

    wc = wb.create_sheet("choices")
    wc.append(CHOICE_COLS)
    for r in choices:
        wc.append([r.get(c, "") for c in CHOICE_COLS])

    wst = wb.create_sheet("settings")
    setting_cols = ["form_title", "form_id", "version", "default_language",
                    "instance_name", "style", "allow_choice_duplicates"]
    wst.append(setting_cols)
    wst.append([
        v["title"], v["form_id"], VERSION_STAMP, v["default_language"],
        "concat(${q1_04_settlement}, '/', ${q1_06_structure}, '/', ${q1_07_hh_serial})",
        "pages", "yes",
    ])
    wb.save(path)


def build_variant(variant: str) -> tuple[int, list[str]]:
    """Write and convert one variant. Returns (status, log lines)."""
    v = VARIANTS[variant]
    base = f"{v['form_id']}_v{VERSION.replace('.', '-')}"
    xlsx = os.path.join(FORM_DIR, base + ".xlsx")
    xml = os.path.join(FORM_DIR, base + ".xml")

    write_xlsx(xlsx, variant)
    print(f"XLSForm written: {xlsx}")
    print(f"  form_id     : {v['form_id']}")
    print(f"  opens in    : {v['default_language']}")
    print(f"  survey rows : {len(survey)}")
    print(f"  choice rows : {len(choices)}")

    import pyxform
    from pyxform.xls2xform import convert

    buf = io.StringIO()
    status = 0
    try:
        with redirect_stdout(buf), redirect_stderr(buf):
            result = convert(xlsx, validate=False, pretty_print=True)
        with open(xml, "w", encoding="utf-8") as fh:
            fh.write(result.xform)
        warnings = list(result.warnings or [])
    except Exception as exc:  # noqa: BLE001 - we want the message in the log
        warnings = []
        status = 1
        buf.write(f"\nCONVERSION FAILED: {type(exc).__name__}: {exc}\n")

    lines = [
        "=" * 78,
        "XLSForm CONVERSION OUTPUT",
        "=" * 78,
        f"tool           : pyxform {pyxform.__version__}",
        f"python         : {sys.version.split()[0]}",
        f"platform target: KoboToolbox (Enketo web forms + KoboCollect / ODK Collect)",
        f"input          : {os.path.basename(xlsx)}",
        f"output         : {os.path.basename(xml)}",
        f"form_id        : {v['form_id']}",
        f"variant        : {variant} ({v['form_variant']}) - opens in {v['default_language']}",
        f"                 {v['note']}",
        f"version        : {VERSION_STAMP}  (semantic {VERSION})",
        f"survey rows    : {len(survey)}",
        f"choice rows    : {len(choices)}",
        "",
        f"RESULT         : {'CONVERTED WITHOUT ERROR' if status == 0 else 'FAILED'}",
        f"warnings       : {len(warnings)}",
        "",
    ]
    if warnings:
        lines.append("-" * 78)
        lines.append("PYXFORM WARNINGS")
        lines.append("-" * 78)
        lines += [f"  [{i + 1}] {w}" for i, w in enumerate(warnings)]
        lines.append("")
    captured = buf.getvalue().strip()
    if captured:
        lines += ["-" * 78, "CAPTURED OUTPUT", "-" * 78, captured, ""]
    lines += [
        "-" * 78,
        "NOTE ON VALIDATION",
        "-" * 78,
        "pyxform's own model check is run above. ODK Validate (the JavaRosa-backed",
        "checker) requires a Java runtime, which is not present on this machine; the",
        "equivalent structural assertions are therefore made directly against the",
        "compiled XForm by scripts/validate_form.py, whose output is in",
        "form/validation_output.txt. Both must pass before deployment.",
        "",
    ]
    return status, lines


def main() -> int:
    os.makedirs(FORM_DIR, exist_ok=True)
    all_lines: list[str] = []
    status = 0
    for variant in VARIANTS:
        st, lines = build_variant(variant)
        status |= st
        all_lines += lines + [""]
    all_lines += [
        "=" * 78,
        "THE TWO VARIANTS ARE THE SAME INSTRUMENT",
        "=" * 78,
        "Both carry both translations and are identical in every question,",
        "constraint, relevance rule and calculation. They differ only in form_id,",
        "form_title, the form_variant value carried on each submission, and which",
        "language the client opens in.",
        "",
        "validate_form.py runs the full structural suite over each of them and",
        "asserts, bind by bind, that the two models are identical apart from the",
        "form_variant calculate. A rule added to one and not the other fails the",
        "build.",
        "",
    ]
    log_path = os.path.join(FORM_DIR, "conversion_output.txt")
    with open(log_path, "w", encoding="utf-8") as fh:
        fh.write("\n".join(all_lines))
    print("\n".join(all_lines))
    return status


if __name__ == "__main__":
    sys.exit(main())
